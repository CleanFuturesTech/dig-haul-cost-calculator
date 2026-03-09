"""
Dig and Haul Cost Calculator - Streamlit Web App v4.4
Run with: streamlit run dig_and_haul_app_v4.4.py

Version 2.0: Updated equipment productivity defaults to medium-class raw CY/hr midpoints;
             added full plain-text report export (assumptions + results) for AI chat use
Version 2.1-2.9: Various defaults, backfill %, report methodology, paid/productive hours,
             daily volume cap, truck logic fixes, trip rounding
Version 3.0: Standard trip rounding (.5+ up, .4- down)
Version 3.1: Major cost model restructure —
             - Equipment charged at daily rate (separate from operator)
             - Operators charged hourly with 1.5x OT on weekend days
             - Work days per week input (5/6/7, default 5)
             - Mob/Demob charge per unit, separate for excavators and loaders
             - Crew truck added (per-day charge, not billed on weather days)
             - Inclement weather days input (extends calendar, equipment still billed,
               operators and crew truck not billed)
             - Environmental Compliance & Insurance fee (% of equip+operators+crew truck)
             - Energy Surcharge (% of heavy equipment cost)
             - Environmental Consulting ($/CY)
             - Site Access Construction Contingency (flat $)
Version 3.2: Operator OT logic revised — OT now based on 40-hr weekly threshold
             rather than weekend-day detection. Added Operator Paid Hours/Day input
             (default 10, yard-to-yard). Partial weeks handled precisely: hours
             accumulate day by day within the partial week, OT kicks in after hr 40.
Version 3.3: No calculation changes — clarity update only. Improved sidebar help
             text and captions to explain the three-way hours split (productive hrs,
             operator paid hrs, daily equipment rate). Added live hours summary
             caption. Report Section 3 now has a dedicated "Three Types of Hours"
             block with worked examples.
Version 3.4: Updated default values to match project-specific inputs. Trucking
             cost model changed from trip-based to time-based — trucks contracted
             for the project are paid for all productive hours on site, so adding
             excess trucks now correctly increases cost. Truck utilization % added
             to Capacity tab and report.
Version 3.5: Fixed calendar days calculation — was incorrectly equal to working
             days (only added weather days). Now correctly accounts for weekends:
             Calendar Days = (Complete Weeks x 7) + Remaining Days + Weather Days.
Version 3.6: Trucking cost corrected — now uses operator_paid_hours_per_day (trucks
             operate in productive hours but are paid for full yard-to-yard day).
             Truck hourly rate default updated to $105. Crew & Site section renamed
             to Miscellaneous Equipment; added Porta Potty, Safety Trailer, and
             Dump Trailer as $/day inputs (default $0). EC&I base updated to
             include all misc equipment.
Version 3.7: Fixed circular truck utilization recommendation — was using
             floor(num_trucks × util%) which chased itself lower each time a
             truck was removed. Now uses ceil(excavation_CY/day /
             (trips_per_truck × truck_capacity)), stable and independent of
             current truck count.
Version 3.8: Mob/Demob inputs moved to Miscellaneous Equipment section; defaults
             lowered to $2,500/unit. Backfill at Landfill unchecked by default.
             Backfill travel field renamed to "Additional Travel Time for Backfill"
             to clarify it is additive to the standard round-trip cycle time.
"""

import streamlit as st
import pandas as pd
import math
import os
import io
from pathlib import Path
from datetime import date

# ReportLab — for PDF quote generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    HRFlowable, KeepTogether, Image, PageBreak
)
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT

APP_VERSION = "4.4"

# ── Page config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Dig and Haul Cost Calculator",
    page_icon="🚜",
    layout="wide"
)

st.markdown("""
    <style>
    .big-font { font-size:20px !important; font-weight: bold; }
    .metric-card { background-color: #f0f2f6; padding: 20px;
                   border-radius: 10px; margin: 10px 0; }
    </style>
    """, unsafe_allow_html=True)

# ── Header ───────────────────────────────────────────────────────────────────
col1, col2, col3 = st.columns([1, 2, 1])
with col2:
    logo_path = Path("Clean_Futures_2.png")
    if logo_path.exists():
        st.image(str(logo_path), use_container_width=True)
    else:
        st.markdown("<h2 style='text-align: center;'>Clean Futures</h2>",
                    unsafe_allow_html=True)
    st.markdown("<h1 style='text-align: center;'>Dig and Haul Cost Calculator</h1>",
                unsafe_allow_html=True)
    st.markdown("<p style='text-align: center;'>Calculate costs and CO2 emissions for "
                "excavating contaminated soil and replacing with clean backfill</p>",
                unsafe_allow_html=True)
    st.markdown(f"<p style='text-align: center; color: gray; font-size: 13px;'>"
                f"Version {APP_VERSION}</p>", unsafe_allow_html=True)

# ── Sidebar ───────────────────────────────────────────────────────────────────
st.sidebar.header("📋 Project Inputs")

# Project Information
st.sidebar.subheader("Project Information")
total_volume = st.sidebar.number_input(
    "Total Volume to Excavate (CY)", min_value=1, value=87000, step=50)
backfill_pct = st.sidebar.number_input(
    "Backfill Volume (% of excavated)", min_value=0, max_value=100, value=100, step=5)
backfill_volume = total_volume * (backfill_pct / 100)
st.sidebar.caption(f"Backfill volume: {backfill_volume:,.0f} CY "
                   f"({backfill_pct}% of {total_volume:,} CY)")
productive_hours_per_day = st.sidebar.number_input(
    "Productive Hours per Day", min_value=1, value=8, step=1,
    help="Hours the equipment is actively running and moving soil. "
         "Drives: excavation volume/day, truck trips/day, and CO2 calculations. "
         "Typically less than paid hours due to yard travel, breaks, and inspections.")
operator_paid_hours_per_day = st.sidebar.number_input(
    "Operator Paid Hours per Day", min_value=1, value=10, step=1,
    help="Yard-to-yard hours — operators are paid from the time they leave the yard "
         "to the time they return. Typically longer than productive hours. "
         "Drives: operator cost and OT threshold only. Does NOT affect volume or truck trips.")
work_days_per_week = st.sidebar.selectbox(
    "Work Days per Week", options=[5, 6, 7], index=0,
    help="Number of days worked per week. OT is calculated on a 40-hr weekly threshold — "
         "any hours above 40 in a week are billed at 1.5x regardless of which day they fall on.")
st.sidebar.caption(
    f"📌 **Hours summary:** Equipment runs productively for {productive_hours_per_day} hrs/day "
    f"(volume & truck trips). Operators are paid for {operator_paid_hours_per_day} hrs/day "
    f"(yard-to-yard). Heavy equipment billed at a flat daily rate.")
weather_days = st.sidebar.number_input(
    "Inclement Weather Days", min_value=0, value=0, step=1,
    help="Days lost to weather. Equipment is still billed at daily rate. "
         "Operators and crew truck are NOT billed. Extends calendar duration.")

# Excavation Equipment
st.sidebar.subheader("Excavation Equipment")
num_excavators = st.sidebar.number_input(
    "Number of Excavators", min_value=0, value=1, step=1)
excavator_daily_rate = st.sidebar.number_input(
    "Excavator Daily Rate ($/day)", min_value=0, value=550, step=50,
    help="Bare equipment rental rate per day, including fuel. "
         "Billed for all project days AND weather days.")
excavator_operator_rate = st.sidebar.number_input(
    "Excavator Operator Rate ($/hr)", min_value=0, value=65, step=5,
    help="Operator hourly rate. 1.5x applied on weekend days if working 6 or 7 days/week.")
excavator_fuel = st.sidebar.number_input(
    "Excavator Fuel (gal/hr) — CO2 tracking", min_value=0.0, value=6.0, step=0.5)
excavator_capacity = st.sidebar.number_input(
    "Excavator Production (CY/hr)", min_value=0, value=105, step=5)

# Loader Equipment
st.sidebar.subheader("Loader Equipment")
num_loaders = st.sidebar.number_input(
    "Number of Loaders", min_value=0, value=1, step=1)
loader_daily_rate = st.sidebar.number_input(
    "Loader Daily Rate ($/day)", min_value=0, value=415, step=50,
    help="Bare equipment rental rate per day, including fuel. "
         "Billed for all project days AND weather days.")
loader_operator_rate = st.sidebar.number_input(
    "Loader Operator Rate ($/hr)", min_value=0, value=65, step=5,
    help="Operator hourly rate. 1.5x applied on weekend days if working 6 or 7 days/week.")
loader_fuel = st.sidebar.number_input(
    "Loader Fuel (gal/hr) — CO2 tracking", min_value=0.0, value=5.0, step=0.5)
loader_capacity = st.sidebar.number_input(
    "Loader Production (CY/hr)", min_value=0, value=130, step=5)

# Daily Volume Cap
max_volume_per_pair = st.sidebar.number_input(
    "Max Daily Volume per Equipment Pair (CY)",
    min_value=0, value=750, step=50,
    help="Real-world ceiling on daily excavation volume per excavator/loader pair.")
st.sidebar.caption("Accounts for logistics, inspections, operator breaks, truck queuing.")

# Trucking
st.sidebar.subheader("Trucking")
num_trucks = st.sidebar.number_input(
    "Number of Trucks", min_value=1, value=5, step=1)
truck_capacity = st.sidebar.number_input(
    "Truck Capacity (CY)", min_value=1, value=18, step=1)
truck_hourly_rate = st.sidebar.number_input(
    "Truck Hourly Rate ($/hr, includes driver & fuel)", min_value=0, value=105, step=5,
    help="Trucks operate (make trips) within productive hours, but are paid for the full "
         "operator paid hours/day. Cost = num trucks × paid hrs/day × rate × project days.")
truck_fuel_rate = st.sidebar.number_input(
    "Truck Fuel (gal/hr) — CO2 tracking", min_value=0.0, value=4.0, step=0.5)

# Miscellaneous Equipment (Daily Charges)
st.sidebar.subheader("Miscellaneous Equipment")
st.sidebar.caption("Daily items billed per working day. Not charged on weather days.")
num_crew_trucks = st.sidebar.number_input(
    "Number of Crew Trucks", min_value=0, value=1, step=1)
crew_truck_daily_rate = st.sidebar.number_input(
    "Crew Truck Daily Rate ($/day)", min_value=0, value=300, step=25,
    help="Billed on working days only — NOT billed on weather days.")
porta_potty_daily_rate = st.sidebar.number_input(
    "Porta Potty ($/day)", min_value=0, value=0, step=5,
    help="Daily rental charge for portable sanitation on site.")
safety_trailer_daily_rate = st.sidebar.number_input(
    "Safety Trailer ($/day)", min_value=0, value=0, step=25,
    help="Daily rental charge for on-site safety trailer.")
dump_trailer_daily_rate = st.sidebar.number_input(
    "Dump Trailer ($/day)", min_value=0, value=0, step=25,
    help="Daily rental charge for dump trailer, if needed.")
num_spotters = st.sidebar.number_input(
    "Number of Spotters", min_value=0, value=0, step=1)
spotter_daily_rate = st.sidebar.number_input(
    "Spotter Daily Rate ($/day)", min_value=0, value=450, step=25,
    help="Billed on working days only.")
num_supervisors = st.sidebar.number_input(
    "Number of Supervisors", min_value=0, value=0, step=1)
supervisor_daily_rate = st.sidebar.number_input(
    "Supervisor Daily Rate ($/day)", min_value=0, value=950, step=25,
    help="Billed on working days only.")
per_diem_daily_rate = st.sidebar.number_input(
    "Per Diems — Flat Daily Rate ($)", min_value=0, value=0, step=25,
    help="Flat daily per diem amount for the full crew (all-in). Multiplied by working days.")
st.sidebar.caption("Mob/Demob — one-way charge per unit, assessed twice (mob + demob).")
excavator_mob_rate = st.sidebar.number_input(
    "Excavator Mob/Demob ($/unit)", min_value=0, value=2500, step=100,
    help="One-way charge per excavator. Assessed twice (mobilization + demobilization).")
loader_mob_rate = st.sidebar.number_input(
    "Loader Mob/Demob ($/unit)", min_value=0, value=2500, step=100,
    help="One-way charge per loader. Assessed twice (mobilization + demobilization).")

# Fees & Contingencies
st.sidebar.subheader("Fees & Contingencies")
eci_pct = st.sidebar.number_input(
    "Environmental Compliance & Insurance (%)", min_value=0.0, value=12.0, step=0.5,
    help="Applied as % of (heavy equipment + operators + crew truck).")
energy_surcharge_pct = st.sidebar.number_input(
    "Energy Surcharge (%)", min_value=0.0, value=28.0, step=0.5,
    help="Applied as % of heavy equipment cost only.")
env_consulting_rate = st.sidebar.number_input(
    "Environmental Consulting ($/day)", min_value=0, value=1500, step=50,
    help="Daily rate for the environmental consultant. Multiplied by Days Onsite below.")
env_consulting_days = st.sidebar.number_input(
    "Environmental Consulting — Days Onsite", min_value=0, value=0, step=1,
    help="Number of days the consultant is on site. Does not need to equal project working days.")
site_access_contingency = st.sidebar.number_input(
    "Site Access Construction Contingency ($)", min_value=0, value=0, step=500,
    help="Flat dollar amount for roadwork or access construction, if required.")

# Fuel Surcharge
st.sidebar.subheader("Fuel Surcharge (Optional)")
fuel_surcharge_enabled = st.sidebar.checkbox("Enable Fuel Surcharge", value=False)
if fuel_surcharge_enabled:
    fuel_surcharge_amount = st.sidebar.number_input(
        "Surcharge Amount ($)", min_value=0, value=250, step=50)
    fuel_surcharge_interval = st.sidebar.selectbox(
        "Surcharge Interval", ["daily", "weekly", "per-trip"])
else:
    fuel_surcharge_amount = 0
    fuel_surcharge_interval = "daily"

# Trip Times
st.sidebar.subheader("Trip Times")
loading_time = st.sidebar.number_input(
    "Loading Time (hours)", min_value=0.0, value=0.10, step=0.05)
travel_time = st.sidebar.number_input(
    "Travel to Landfill (hours, one-way)", min_value=0.0, value=0.5, step=0.1)
landfill_time = st.sidebar.number_input(
    "Time at Landfill (wait + dump, hours)", min_value=0.0, value=0.25, step=0.05)

# Backfill
st.sidebar.subheader("Backfill")
backfill_at_landfill = st.sidebar.checkbox("Backfill Available at Landfill", value=False)
backfill_cost = st.sidebar.number_input(
    "Backfill Cost ($/CY)", min_value=0, value=10, step=1)
if not backfill_at_landfill:
    travel_to_backfill = st.sidebar.number_input(
        "Additional Travel Time for Backfill (hours)", min_value=0.0, value=0.25, step=0.05,
        help="Extra time added to the round-trip cycle for backfill pickup. "
             "Added on top of the standard landfill round-trip time.")
    backfill_loading_time = st.sidebar.number_input(
        "Backfill Loading Time (hours)", min_value=0.0, value=0.10, step=0.05)
else:
    travel_to_backfill = 0
    backfill_loading_time = 0

# Disposal
st.sidebar.subheader("Disposal")
disposal_cost = st.sidebar.number_input(
    "Disposal Cost ($/CY)", min_value=0, value=25, step=1)

# Calculate button
calculate = st.sidebar.button("🧮 Calculate", type="primary")

# ── Calculations ──────────────────────────────────────────────────────────────
if calculate or 'results' in st.session_state:

    # ── Excavation capacity ──
    excavator_total_capacity = num_excavators * excavator_capacity
    loader_total_capacity    = num_loaders    * loader_capacity

    if excavator_total_capacity > 0 and loader_total_capacity > 0:
        excavation_capacity = min(excavator_total_capacity, loader_total_capacity)
        if   excavator_total_capacity < loader_total_capacity: equipment_bottleneck = "Excavator"
        elif loader_total_capacity < excavator_total_capacity: equipment_bottleneck = "Loader"
        else:                                                   equipment_bottleneck = "Balanced"
        num_pairs = min(num_excavators, num_loaders)
    elif excavator_total_capacity > 0:
        excavation_capacity  = excavator_total_capacity
        equipment_bottleneck = "N/A"
        num_pairs            = num_excavators
    else:
        excavation_capacity  = loader_total_capacity
        equipment_bottleneck = "N/A"
        num_pairs            = num_loaders

    excavation_volume_per_day_uncapped = excavation_capacity * productive_hours_per_day
    daily_volume_cap = (num_pairs * max_volume_per_pair
                        if max_volume_per_pair > 0
                        else excavation_volume_per_day_uncapped)
    excavation_volume_per_day = min(excavation_volume_per_day_uncapped, daily_volume_cap)
    volume_cap_active = excavation_volume_per_day < excavation_volume_per_day_uncapped

    # ── Trip time ──
    if backfill_at_landfill:
        trip_time = loading_time + travel_time + landfill_time + travel_time + loading_time
    else:
        trip_time = (loading_time + travel_time + landfill_time
                     + travel_to_backfill + backfill_loading_time
                     + travel_time + loading_time)

    # ── Trucking capacity (no rounding — actual fractional trips) ──
    trips_per_truck_per_day          = productive_hours_per_day / trip_time
    trips_per_truck_per_day_raw      = trips_per_truck_per_day   # kept for compatibility
    total_trips_per_day_theoretical  = trips_per_truck_per_day * num_trucks
    total_trips_per_day_theoretical_raw = total_trips_per_day_theoretical
    truck_volume_per_day_theoretical = total_trips_per_day_theoretical * truck_capacity

    truck_volume_per_day      = min(truck_volume_per_day_theoretical, excavation_volume_per_day)
    effective_trips_per_day   = truck_volume_per_day / truck_capacity
    effective_trips_per_day_raw = effective_trips_per_day

    # ── Bottleneck & project duration ──
    limiting_volume = min(excavation_volume_per_day, truck_volume_per_day_theoretical)
    bottleneck = "Trucking" if truck_volume_per_day_theoretical <= excavation_volume_per_day else "Excavation"

    project_days    = math.ceil(total_volume / limiting_volume)
    # Calendar days = working days expanded to full weeks (including weekends) + weather days
    _complete_weeks  = project_days // work_days_per_week
    _remaining_days  = project_days  % work_days_per_week
    calendar_days   = _complete_weeks * 7 + _remaining_days + weather_days
    num_trips       = math.ceil(total_volume / truck_capacity)

    # ── Operator OT — 40-hr weekly threshold ──
    # Operators paid yard-to-yard (operator_paid_hours_per_day).
    # Any hours over 40 in a calendar week are OT at 1.5x, regardless of day.
    weekly_paid_hours      = operator_paid_hours_per_day * work_days_per_week
    regular_hours_per_week = min(40, weekly_paid_hours)
    ot_hours_per_week      = max(0, weekly_paid_hours - 40)

    # Split project into complete weeks + partial-week remainder
    complete_weeks = project_days // work_days_per_week
    remaining_days = project_days  % work_days_per_week

    # Partial week: hours accumulate; OT kicks in after hr 40
    partial_hours         = remaining_days * operator_paid_hours_per_day
    partial_regular_hours = min(40, partial_hours)
    partial_ot_hours      = max(0, partial_hours - 40)

    # Helper: total project cost for one operator at a given rate
    def operator_project_cost(rate):
        complete = complete_weeks * (
            regular_hours_per_week * rate +
            ot_hours_per_week      * rate * 1.5)
        partial  = (partial_regular_hours * rate +
                    partial_ot_hours      * rate * 1.5)
        return complete + partial

    # OT summary values for display / report
    total_operator_regular_hrs = (complete_weeks * regular_hours_per_week
                                  + partial_regular_hours)
    total_operator_ot_hrs      = (complete_weeks * ot_hours_per_week
                                  + partial_ot_hours)

    # ── Cost calculations ──

    # 1. Heavy equipment — daily rate × (project days + weather days)
    total_billed_equipment_days = project_days + weather_days
    excavator_equipment_cost = num_excavators * excavator_daily_rate * total_billed_equipment_days
    loader_equipment_cost    = num_loaders    * loader_daily_rate    * total_billed_equipment_days
    total_equipment_cost     = excavator_equipment_cost + loader_equipment_cost

    # 2. Mob/Demob — 2× per unit (mob + demob)
    excavator_mob_cost = num_excavators * excavator_mob_rate * 2
    loader_mob_cost    = num_loaders    * loader_mob_rate    * 2
    total_mob_cost     = excavator_mob_cost + loader_mob_cost

    # 3. Operators — 40-hr OT threshold; not paid on weather days
    excavator_operator_cost = num_excavators * operator_project_cost(excavator_operator_rate)
    loader_operator_cost    = num_loaders    * operator_project_cost(loader_operator_rate)
    total_operator_cost     = excavator_operator_cost + loader_operator_cost

    # 4. Misc equipment — daily × project_days only (not weather days)
    crew_truck_cost        = num_crew_trucks * crew_truck_daily_rate * project_days
    porta_potty_cost       = porta_potty_daily_rate * project_days
    safety_trailer_cost    = safety_trailer_daily_rate * project_days
    dump_trailer_cost      = dump_trailer_daily_rate * project_days
    spotter_cost           = num_spotters * spotter_daily_rate * project_days
    supervisor_cost        = num_supervisors * supervisor_daily_rate * project_days
    per_diem_cost          = per_diem_daily_rate * project_days
    total_misc_cost        = (crew_truck_cost + porta_potty_cost + safety_trailer_cost +
                              dump_trailer_cost + spotter_cost + supervisor_cost + per_diem_cost)

    # 5. Energy Surcharge — % of heavy equipment
    energy_surcharge_cost = (energy_surcharge_pct / 100) * total_equipment_cost

    # 6. Environmental Compliance & Insurance — % of (equipment + operators + misc equipment)
    eci_base = total_equipment_cost + total_operator_cost + total_misc_cost
    eci_cost = (eci_pct / 100) * eci_base

    # 7. Trucking — trucks are on-site and on the clock for the full productive day.
    #    Paying for num_trucks × hours on site × project days, regardless of whether
    #    they are running or waiting on excavation.
    #    num_trips is kept for disposal cost, fuel surcharge, and trip display.
    total_truck_hours = num_trips * trip_time   # kept for CO2 / surcharge / display
    # Trucks operate (make trips) within productive hours, but are paid for full paid hours/day
    trucking_cost = num_trucks * operator_paid_hours_per_day * truck_hourly_rate * project_days
    # Utilization: active trip hours vs. total contracted truck hours (productive only)
    active_truck_hours = effective_trips_per_day * trip_time * project_days
    total_contracted_truck_hours = num_trucks * operator_paid_hours_per_day * project_days
    truck_utilization_pct = (active_truck_hours / total_contracted_truck_hours * 100
                             if total_contracted_truck_hours > 0 else 0)

    # 8. Fuel Surcharge
    if fuel_surcharge_enabled:
        if   fuel_surcharge_interval == "daily":
            fuel_surcharge_cost = fuel_surcharge_amount * project_days
        elif fuel_surcharge_interval == "weekly":
            fuel_surcharge_cost = fuel_surcharge_amount * math.ceil(project_days / 7)
        elif fuel_surcharge_interval == "per-trip":
            fuel_surcharge_cost = fuel_surcharge_amount * num_trips
        else:
            fuel_surcharge_cost = 0
    else:
        fuel_surcharge_cost = 0

    # 9. Disposal & Backfill
    total_disposal_cost = total_volume  * disposal_cost
    total_backfill_cost = backfill_volume * backfill_cost

    # 10. Environmental Consulting — daily rate × days onsite
    env_consulting_cost = env_consulting_rate * env_consulting_days

    # 11. Site Access Contingency (flat)
    # already a dollar value

    # ── Grand total ──
    total_cost = (
        total_equipment_cost +
        total_mob_cost +
        total_operator_cost +
        total_misc_cost +
        energy_surcharge_cost +
        eci_cost +
        trucking_cost +
        fuel_surcharge_cost +
        total_disposal_cost +
        total_backfill_cost +
        env_consulting_cost +
        site_access_contingency
    )
    cost_per_cy = total_cost / total_volume

    # ── CO2 (productive hours only) ──
    productive_project_hours = project_days * productive_hours_per_day
    total_fuel_gallons = (
        num_excavators * excavator_fuel * productive_project_hours +
        num_loaders    * loader_fuel    * productive_project_hours +
        total_truck_hours * truck_fuel_rate)
    co2_lbs  = total_fuel_gallons * 22.4
    co2_tons = co2_lbs / 2000

    # ── Store results ──
    st.session_state['results'] = {
        # Schedule
        'project_days':   project_days,
        'calendar_days':  calendar_days,
        'weather_days':   weather_days,
        'num_trips':      num_trips,
        'bottleneck':     bottleneck,
        # Capacity
        'excavator_capacity':               excavator_total_capacity,
        'loader_capacity':                  loader_total_capacity,
        'excavation_capacity':              excavation_capacity,
        'equipment_bottleneck':             equipment_bottleneck,
        'num_pairs':                        num_pairs,
        'excavation_volume_per_day_uncapped': excavation_volume_per_day_uncapped,
        'daily_volume_cap':                 daily_volume_cap,
        'volume_cap_active':                volume_cap_active,
        'excavation_volume_per_day':        excavation_volume_per_day,
        'truck_volume_per_day':             truck_volume_per_day,
        'truck_volume_per_day_theoretical': truck_volume_per_day_theoretical,
        'effective_trips_per_day':          effective_trips_per_day,
        'effective_trips_per_day_raw':      effective_trips_per_day_raw,
        'trips_per_truck_per_day':          trips_per_truck_per_day,
        'trips_per_truck_per_day_raw':      trips_per_truck_per_day_raw,
        'total_trips_per_day_theoretical':  total_trips_per_day_theoretical,
        'total_trips_per_day_theoretical_raw': total_trips_per_day_theoretical_raw,
        'trip_time':                        trip_time,
        # OT summary
        'weekly_paid_hours':         weekly_paid_hours,
        'regular_hours_per_week':    regular_hours_per_week,
        'ot_hours_per_week':         ot_hours_per_week,
        'complete_weeks':            complete_weeks,
        'remaining_days':            remaining_days,
        'total_operator_regular_hrs': total_operator_regular_hrs,
        'total_operator_ot_hrs':     total_operator_ot_hrs,
        # Costs
        'total_equipment_cost':     total_equipment_cost,
        'excavator_equipment_cost': excavator_equipment_cost,
        'loader_equipment_cost':    loader_equipment_cost,
        'total_mob_cost':           total_mob_cost,
        'excavator_mob_cost':       excavator_mob_cost,
        'loader_mob_cost':          loader_mob_cost,
        'total_operator_cost':      total_operator_cost,
        'excavator_operator_cost':  excavator_operator_cost,
        'loader_operator_cost':     loader_operator_cost,
        'crew_truck_cost':          crew_truck_cost,
        'porta_potty_cost':         porta_potty_cost,
        'safety_trailer_cost':      safety_trailer_cost,
        'dump_trailer_cost':        dump_trailer_cost,
        'spotter_cost':             spotter_cost,
        'supervisor_cost':          supervisor_cost,
        'per_diem_cost':            per_diem_cost,
        'total_misc_cost':          total_misc_cost,
        'energy_surcharge_cost':    energy_surcharge_cost,
        'eci_cost':                 eci_cost,
        'trucking_cost':            trucking_cost,
        'total_truck_hours':        total_truck_hours,
        'active_truck_hours':       active_truck_hours,
        'total_contracted_truck_hours': total_contracted_truck_hours,
        'truck_utilization_pct':    truck_utilization_pct,
        'fuel_surcharge_cost':      fuel_surcharge_cost,
        'total_disposal_cost':      total_disposal_cost,
        'total_backfill_cost':      total_backfill_cost,
        'env_consulting_cost':      env_consulting_cost,
        'site_access_contingency':  site_access_contingency,
        'total_cost':               total_cost,
        'cost_per_cy':              cost_per_cy,
        # CO2
        'total_fuel_gallons': total_fuel_gallons,
        'co2_tons':           co2_tons,
    }

    results = st.session_state['results']

    # ── Recalculate button ──
    recalc_col, _ = st.columns([1, 4])
    with recalc_col:
        st.button("🔄 Recalculate", type="primary", key="recalc_btn")

    # ── Results summary metrics ──
    st.header("📊 Results Summary")
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Total Cost",    f"${results['total_cost']:,.0f}")
        st.metric("Cost per CY",   f"${results['cost_per_cy']:.2f}")
    with col2:
        st.metric("Working Days",  f"{results['project_days']} days")
        st.metric("Calendar Days", f"{results['calendar_days']} days"
                  + (f" (+{weather_days} weather)" if weather_days > 0 else ""))
    with col3:
        st.metric("CO2 Emissions", f"{results['co2_tons']:.2f} tons")
        st.metric("Truck Trips",   f"{results['num_trips']} trips")
    with col4:
        st.metric("Bottleneck",       results['bottleneck'])
        st.metric("Equipment Limit",  results['equipment_bottleneck'])

    # ── Detailed analysis tabs ──
    st.header("📈 Detailed Analysis")
    tab1, tab2, tab3 = st.tabs(
        ["💰 Cost Breakdown", "⚙️ Capacity Analysis", "🌍 Environmental Impact"])

    # ── Tab 1: Cost Breakdown ──
    with tab1:
        st.subheader("Cost Breakdown")
        col1, col2 = st.columns(2)

        with col1:
            cost_data = {
                'Category': [
                    'Heavy Equipment (daily)',
                    '  — Excavators',
                    '  — Loaders',
                    'Mob / Demob',
                    'Operators',
                    '  — Excavator Operators',
                    '  — Loader Operators',
                    'Misc Equipment',
                    '  — Crew Truck(s)',
                    '  — Porta Potty',
                    '  — Safety Trailer',
                    '  — Dump Trailer',
                    '  — Spotters',
                    '  — Supervisors',
                    '  — Per Diems',
                    'Energy Surcharge',
                    'EC&I Fee',
                    'Trucking',
                    'Fuel Surcharge',
                    'Disposal',
                    'Backfill Material',
                    'Environmental Consulting',
                    'Site Access Contingency',
                ],
                'Cost': [
                    f"${results['total_equipment_cost']:,.0f}",
                    f"  ${results['excavator_equipment_cost']:,.0f}",
                    f"  ${results['loader_equipment_cost']:,.0f}",
                    f"${results['total_mob_cost']:,.0f}",
                    f"${results['total_operator_cost']:,.0f}",
                    f"  ${results['excavator_operator_cost']:,.0f}",
                    f"  ${results['loader_operator_cost']:,.0f}",
                    f"${results['total_misc_cost']:,.0f}",
                    f"  ${results['crew_truck_cost']:,.0f}",
                    f"  ${results['porta_potty_cost']:,.0f}",
                    f"  ${results['safety_trailer_cost']:,.0f}",
                    f"  ${results['dump_trailer_cost']:,.0f}",
                    f"  ${results['spotter_cost']:,.0f}",
                    f"  ${results['supervisor_cost']:,.0f}",
                    f"  ${results['per_diem_cost']:,.0f}",
                    f"${results['energy_surcharge_cost']:,.0f}",
                    f"${results['eci_cost']:,.0f}",
                    f"${results['trucking_cost']:,.0f}",
                    f"${results['fuel_surcharge_cost']:,.0f}",
                    f"${results['total_disposal_cost']:,.0f}",
                    f"${results['total_backfill_cost']:,.0f}",
                    f"${results['env_consulting_cost']:,.0f}",
                    f"${results['site_access_contingency']:,.0f}",
                ]
            }
            df_costs = pd.DataFrame(cost_data)
            st.dataframe(df_costs, hide_index=True, use_container_width=True)
            st.metric("**TOTAL**", f"${results['total_cost']:,.0f}")

            if weather_days > 0:
                st.info(f"⛈️ **Weather note:** Equipment billed for "
                        f"{project_days + weather_days} days "
                        f"({project_days} working + {weather_days} weather). "
                        f"Operators and misc equipment billed for {project_days} working days only.")
            if results['total_operator_ot_hrs'] > 0:
                ot_pct = results['total_operator_ot_hrs'] / (results['total_operator_regular_hrs'] + results['total_operator_ot_hrs']) * 100
                st.info(f"⏱️ **OT note:** {results['weekly_paid_hours']} hrs/week "
                        f"({operator_paid_hours_per_day} hrs/day × {work_days_per_week} days) "
                        f"exceeds 40-hr threshold by {results['ot_hours_per_week']:.0f} hrs/week. "
                        f"Project total: {results['total_operator_regular_hrs']:.0f} regular hrs + "
                        f"{results['total_operator_ot_hrs']:.0f} OT hrs "
                        f"({ot_pct:.0f}% of operator hours at 1.5x).")

        with col2:
            chart_cats = [
                'Equipment', 'Mob/Demob', 'Operators', 'Misc Equipment',
                'Energy Surcharge', 'EC&I', 'Trucking', 'Fuel Surcharge',
                'Disposal', 'Backfill', 'Env. Consulting', 'Site Access'
            ]
            chart_vals = [
                results['total_equipment_cost'], results['total_mob_cost'],
                results['total_operator_cost'],  results['total_misc_cost'],
                results['energy_surcharge_cost'],results['eci_cost'],
                results['trucking_cost'],        results['fuel_surcharge_cost'],
                results['total_disposal_cost'],  results['total_backfill_cost'],
                results['env_consulting_cost'],  results['site_access_contingency'],
            ]
            chart_data = pd.DataFrame({'Category': chart_cats, 'Cost': chart_vals})
            st.bar_chart(chart_data.set_index('Category'))

    # ── Tab 2: Capacity Analysis ──
    with tab2:
        st.subheader("Capacity Analysis")
        col1, col2 = st.columns(2)

        with col1:
            st.markdown("**Equipment Capacity**")
            capacity_data = {
                'Component': [
                    'Excavator Capacity',
                    'Loader Capacity',
                    'Excavation Capacity (bottleneck)',
                    'Equipment Bottleneck',
                    'Equipment Pairs',
                    'Productive Hours per Day',
                    'Theoretical Volume per Day',
                    'Daily Volume Cap (per pair)',
                    'Effective Excavation Volume/Day',
                    'Volume Cap Active?',
                ],
                'Value': [
                    f"{results['excavator_capacity']} CY/hr" if results['excavator_capacity'] > 0 else 'N/A',
                    f"{results['loader_capacity']} CY/hr"    if results['loader_capacity']    > 0 else 'N/A',
                    f"{results['excavation_capacity']} CY/hr",
                    results['equipment_bottleneck'],
                    f"{results['num_pairs']}",
                    f"{productive_hours_per_day} hrs",
                    f"{results['excavation_volume_per_day_uncapped']:.0f} CY",
                    f"{max_volume_per_pair} CY x {results['num_pairs']} pairs = "
                    f"{results['daily_volume_cap']:.0f} CY",
                    f"{results['excavation_volume_per_day']:.0f} CY",
                    "⚠️ Yes — cap is limiting" if results['volume_cap_active']
                    else "No — theoretical rate governs",
                ]
            }
            st.dataframe(pd.DataFrame(capacity_data), hide_index=True, use_container_width=True)

        with col2:
            st.markdown("**Trucking Capacity**")
            truck_data = {
                'Component': [
                    'Number of Trucks',
                    'Truck Capacity',
                    'Trip Time',
                    'Productive Hours per Day',
                    'Trips per Truck per Day',
                    'Total Trips per Day (theoretical)',
                    'Truck Volume per Day (theoretical)',
                    'Effective Trips per Day (excav-limited)',
                    'Effective Truck Volume per Day',
                ],
                'Value': [
                    f"{num_trucks}",
                    f"{truck_capacity} CY",
                    f"{results['trip_time']:.2f} hrs",
                    f"{productive_hours_per_day} hrs",
                    f"{results['trips_per_truck_per_day']:.2f}",
                    f"{results['total_trips_per_day_theoretical']:.2f}",
                    f"{results['truck_volume_per_day_theoretical']:.1f} CY",
                    f"{results['effective_trips_per_day']:.2f}"
                    if results['bottleneck'] == "Excavation"
                    else "N/A — trucking is bottleneck",
                    f"{results['truck_volume_per_day']:.1f} CY",
                ]
            }
            st.dataframe(pd.DataFrame(truck_data), hide_index=True, use_container_width=True)

        util = results['truck_utilization_pct']
        # Optimal trucks = how many trucks needed to keep pace with excavation output
        # Independent of current num_trucks — avoids circular recommendation
        optimal_trucks = math.ceil(excavation_volume_per_day / (trips_per_truck_per_day * truck_capacity)) if trips_per_truck_per_day > 0 else num_trucks
        util_note = (f"\n        - ✅ Trucks fully utilized ({util:.0f}% of productive hours on trips)"
                     if util >= 90
                     else f"\n        - ⚠️ Truck utilization: {util:.0f}% of productive hours on trips — "
                          f"excavation limits throughput to {results['excavation_volume_per_day']:.0f} CY/day. "
                          f"Optimal truck count: ~{optimal_trucks} trucks.")
        cap_note = (f"\n        - ⚠️ Volume cap active: theoretical "
                    f"{results['excavation_volume_per_day_uncapped']:.0f} CY/day "
                    f"capped at {results['daily_volume_cap']:.0f} CY/day"
                    if results['volume_cap_active'] else "")
        truck_note = (f"\n        - ⚠️ Trucks theoretical ({results['truck_volume_per_day_theoretical']:.0f} "
                      f"CY/day) exceeds excavation — limited to "
                      f"{results['truck_volume_per_day']:.0f} CY/day"
                      if results['bottleneck'] == 'Excavation'
                      and results['truck_volume_per_day_theoretical'] > results['excavation_volume_per_day']
                      else "")
        st.info(f"""
        **System Bottleneck: {results['bottleneck']}**

        - Excavation effective: {results['excavation_volume_per_day']:.0f} CY/day{cap_note}
        - Trucks theoretical:   {results['truck_volume_per_day_theoretical']:.0f} CY/day
        - Trucks effective:     {results['truck_volume_per_day']:.0f} CY/day{truck_note}{util_note}
        - Limiting factor:      {limiting_volume:.0f} CY/day
        - Productive hrs/day:   {productive_hours_per_day} hrs | Work days/week: {work_days_per_week}

        {"Consider adding more trucks to increase productivity."
         if results['bottleneck'] == 'Trucking'
         else "Excavation is limiting. Adding more trucks increases cost without increasing output."}
        """)

    # ── Tab 3: Environmental Impact ──
    with tab3:
        st.subheader("Environmental Impact")
        col1, col2 = st.columns(2)
        with col1:
            env_data = {
                'Metric': [
                    'Total Fuel Consumed',
                    'CO2 Emissions (lbs)',
                    'CO2 Emissions (tons)',
                    'CO2 per Cubic Yard',
                ],
                'Value': [
                    f"{results['total_fuel_gallons']:.0f} gallons",
                    f"{co2_lbs:,.0f} lbs",
                    f"{results['co2_tons']:.2f} tons",
                    f"{co2_lbs / total_volume:.2f} lbs/CY",
                ]
            }
            st.dataframe(pd.DataFrame(env_data), hide_index=True, use_container_width=True)
            st.caption("Fuel consumption uses productive hours only — "
                       "equipment is not running during yard transit or weather days.")
        with col2:
            st.metric("🌳 Equivalent Trees Needed", f"{int(results['co2_tons'] * 16.5):,}")
            st.caption("Trees needed to offset CO2 over 1 year (EPA estimate)")
            st.metric("🚗 Equivalent Car Miles", f"{int(results['co2_tons'] * 2500):,}")
            st.caption("Miles driven by average car (EPA estimate)")

    # ── Download Results ──
    st.header("💾 Download Results")

    # Build text report
    backfill_location_str = ("At landfill (no separate trip)"
                             if backfill_at_landfill else "Separate backfill site")
    backfill_trip_str     = ("N/A (backfill at landfill)" if backfill_at_landfill
                             else f"{travel_to_backfill:.2f} hrs additional travel + "
                                  f"{backfill_loading_time:.2f} hrs loading")
    surcharge_str         = (f"${fuel_surcharge_amount:,} {fuel_surcharge_interval}"
                             if fuel_surcharge_enabled else "Disabled")

    report_lines = [
        "=" * 65,
        "  DIG AND HAUL COST ESTIMATE REPORT",
        f"  Clean Futures | Dig and Haul Cost Calculator v{APP_VERSION}",
        f"  Generated: {date.today().strftime('%B %d, %Y')}",
        "=" * 65,
        "",
        "--- SECTION 1: INPUT ASSUMPTIONS ---",
        "",
        "[ Project Parameters ]",
        f"  Total Volume to Excavate:         {total_volume:,} CY",
        f"  Backfill Volume:                  {backfill_volume:,.0f} CY ({backfill_pct}%)",
        f"  Productive Hours per Day:         {productive_hours_per_day} hrs",
        f"    → drives: excavation volume/day, truck trips/day, CO2",
        f"  Operator Paid Hours per Day:      {operator_paid_hours_per_day} hrs (yard-to-yard)",
        f"    → drives: operator cost and OT threshold only",
        f"  Heavy Equipment:                  flat daily rate (not hour-based)",
        f"  Work Days per Week:               {work_days_per_week} days",
        f"  Weekly Operator Hours:            {weekly_paid_hours} hrs "
        f"({regular_hours_per_week:.0f} regular + {ot_hours_per_week:.0f} OT)",
        f"  Inclement Weather Days:           {weather_days} days",
        "",
        "[ Excavation Equipment ]",
        f"  Number of Excavators:             {num_excavators}",
        f"  Excavator Daily Rate:             ${excavator_daily_rate:,}/day",
        f"  Excavator Operator Rate:          ${excavator_operator_rate}/hr",
        f"  Excavator Mob/Demob:              ${excavator_mob_rate:,}/unit (x2)",
        f"  Excavator Fuel:                   {excavator_fuel} gal/hr (CO2 only)",
        f"  Excavator Production:             {excavator_capacity} CY/hr",
        "",
        "[ Loader Equipment ]",
        f"  Number of Loaders:                {num_loaders}",
        f"  Loader Daily Rate:                ${loader_daily_rate:,}/day",
        f"  Loader Operator Rate:             ${loader_operator_rate}/hr",
        f"  Loader Mob/Demob:                 ${loader_mob_rate:,}/unit (x2)",
        f"  Loader Fuel:                      {loader_fuel} gal/hr (CO2 only)",
        f"  Loader Production:                {loader_capacity} CY/hr",
        "",
        "[ Equipment Productivity Constraints ]",
        f"  Equipment Pairs:                  {results['num_pairs']}",
        f"  Max Daily Volume per Pair:        {max_volume_per_pair} CY",
        f"  Total Daily Volume Cap:           {results['daily_volume_cap']:.0f} CY",
        f"  Theoretical Daily Volume:         {results['excavation_volume_per_day_uncapped']:.0f} CY",
        f"  Effective Daily Volume:           {results['excavation_volume_per_day']:.0f} CY",
        f"  Volume Cap Active:                {'Yes' if results['volume_cap_active'] else 'No'}",
        "",
        "[ Trucking ]",
        f"  Number of Trucks:                 {num_trucks}",
        f"  Truck Capacity:                   {truck_capacity} CY",
        f"  Truck Hourly Rate:                ${truck_hourly_rate}/hr (driver & fuel incl.)",
        f"  Truck Fuel:                       {truck_fuel_rate} gal/hr (CO2 only)",
        "",
        "[ Trip Times ]",
        f"  Loading Time:                     {loading_time:.2f} hrs",
        f"  Travel to Landfill (one-way):     {travel_time:.2f} hrs",
        f"  Time at Landfill:                 {landfill_time:.2f} hrs",
        f"  Full Round-Trip Cycle Time:       {results['trip_time']:.2f} hrs",
        "",
        "[ Miscellaneous Equipment ]",
        f"  Number of Crew Trucks:            {num_crew_trucks}",
        f"  Crew Truck Daily Rate:            ${crew_truck_daily_rate}/day",
        f"  Porta Potty:                      ${porta_potty_daily_rate}/day",
        f"  Safety Trailer:                   ${safety_trailer_daily_rate}/day",
        f"  Dump Trailer:                     ${dump_trailer_daily_rate}/day",
        f"  Number of Spotters:               {num_spotters}",
        f"  Spotter Daily Rate:               ${spotter_daily_rate}/day",
        f"  Number of Supervisors:            {num_supervisors}",
        f"  Supervisor Daily Rate:            ${supervisor_daily_rate}/day",
        f"  Per Diems Flat Daily Rate:        ${per_diem_daily_rate}/day",
        f"  (All misc items billed on working days only — not on weather days)",
        "",
        "[ Fees & Contingencies ]",
        f"  Environmental Compliance & Ins.:  {eci_pct}% of equip + operators + misc equipment",
        f"  Energy Surcharge:                 {energy_surcharge_pct}% of heavy equipment",
        f"  Environmental Consulting:         ${env_consulting_rate}/day × {env_consulting_days} days onsite",
        f"  Site Access Contingency:          ${site_access_contingency:,}",
        "",
        "[ Backfill ]",
        f"  Backfill Location:                {backfill_location_str}",
        f"  Backfill Cost:                    ${backfill_cost}/CY",
        f"  Backfill Additional Travel/Loading: {backfill_trip_str}",
        "",
        "[ Disposal ]",
        f"  Disposal Cost:                    ${disposal_cost}/CY",
        "",
        "[ Fuel Surcharge ]",
        f"  Fuel Surcharge:                   {surcharge_str}",
        "",
        "=" * 65,
        "--- SECTION 2: CALCULATED RESULTS ---",
        "",
        "[ Schedule ]",
        f"  Working Days:                     {results['project_days']} days",
        f"  Weather Days:                     {results['weather_days']} days",
        f"  Total Calendar Days:              {results['calendar_days']} days",
        f"  Complete Weeks:                   {results['complete_weeks']}",
        f"  Remaining Days (partial week):    {results['remaining_days']}",
        f"  Total Truck Trips:                {results['num_trips']:,} trips",
        f"  Trips per Truck per Day:          {results['trips_per_truck_per_day']:.2f}",
        f"  Total Operator Regular Hrs:       {results['total_operator_regular_hrs']:.0f} hrs",
        f"  Total Operator OT Hrs:            {results['total_operator_ot_hrs']:.0f} hrs",
        "",
        "[ Capacity & Bottleneck ]",
        f"  Excavation Capacity (net):        {results['excavation_capacity']} CY/hr",
        f"  Equipment Bottleneck:             {results['equipment_bottleneck']}",
        f"  Effective Excavation Vol/Day:     {results['excavation_volume_per_day']:.0f} CY",
        f"  Effective Truck Vol/Day:          {results['truck_volume_per_day']:.0f} CY",
        f"  System Bottleneck:                {results['bottleneck']}",
        "",
        "[ Cost Breakdown ]",
        f"  Heavy Equipment (daily):          ${results['total_equipment_cost']:>12,.2f}",
        f"    Excavators ({num_excavators} x ${excavator_daily_rate}/day x "
        f"{total_billed_equipment_days} days): ${results['excavator_equipment_cost']:>10,.2f}",
        f"    Loaders    ({num_loaders} x ${loader_daily_rate}/day x "
        f"{total_billed_equipment_days} days): ${results['loader_equipment_cost']:>10,.2f}",
        f"  Mob/Demob:                        ${results['total_mob_cost']:>12,.2f}",
        f"    Excavators ({num_excavators} x ${excavator_mob_rate:,} x 2): "
        f"${results['excavator_mob_cost']:>10,.2f}",
        f"    Loaders    ({num_loaders} x ${loader_mob_rate:,} x 2): "
        f"${results['loader_mob_cost']:>10,.2f}",
        f"  Operators:                        ${results['total_operator_cost']:>12,.2f}",
        f"    Excavator operators:            ${results['excavator_operator_cost']:>12,.2f}",
        f"    Loader operators:               ${results['loader_operator_cost']:>12,.2f}",
        f"  Misc Equipment:                   ${results['total_misc_cost']:>12,.2f}",
        f"    Crew Truck(s):                  ${results['crew_truck_cost']:>12,.2f}",
        f"    Porta Potty:                    ${results['porta_potty_cost']:>12,.2f}",
        f"    Safety Trailer:                 ${results['safety_trailer_cost']:>12,.2f}",
        f"    Dump Trailer:                   ${results['dump_trailer_cost']:>12,.2f}",
        f"    Spotters ({num_spotters} × ${spotter_daily_rate}/day):          "
        f"${results['spotter_cost']:>12,.2f}",
        f"    Supervisors ({num_supervisors} × ${supervisor_daily_rate}/day):  "
        f"${results['supervisor_cost']:>12,.2f}",
        f"    Per Diems (flat ${per_diem_daily_rate}/day):       "
        f"${results['per_diem_cost']:>12,.2f}",
        f"  Energy Surcharge ({energy_surcharge_pct}%):       "
        f"${results['energy_surcharge_cost']:>12,.2f}",
        f"  EC&I Fee ({eci_pct}%):                  ${results['eci_cost']:>12,.2f}",
        f"  Trucking ({num_trucks} trucks × {operator_paid_hours_per_day} paid hrs × "
        f"${truck_hourly_rate}/hr × {results['project_days']} days): "
        f"${results['trucking_cost']:>12,.2f}  [{results['truck_utilization_pct']:.0f}% trip utilization]",
        f"  Fuel Surcharge:                   ${results['fuel_surcharge_cost']:>12,.2f}",
        f"  Disposal:                         ${results['total_disposal_cost']:>12,.2f}",
        f"  Backfill Material:                ${results['total_backfill_cost']:>12,.2f}",
        f"  Environmental Consulting          ${results['env_consulting_cost']:>12,.2f}",
        f"    (${env_consulting_rate}/day × {env_consulting_days} days onsite)",
        f"  Site Access Contingency:          ${results['site_access_contingency']:>12,.2f}",
        f"  {'─' * 45}",
        f"  TOTAL PROJECT COST:               ${results['total_cost']:>12,.2f}",
        f"  Cost per Cubic Yard:              ${results['cost_per_cy']:>12,.2f}",
        "",
        "[ Environmental Impact ]",
        f"  Total Fuel Consumed:              {results['total_fuel_gallons']:,.0f} gallons",
        f"  CO2 Emissions:                    {results['co2_tons']:.2f} tons",
        f"  Equivalent Trees to Offset:       {int(results['co2_tons'] * 16.5):,} (1-year)",
        f"  Equivalent Car Miles:             {int(results['co2_tons'] * 2500):,} miles",
        "",
        "=" * 65,
        "--- SECTION 3: HOW KEY METRICS WERE DERIVED ---",
        "=" * 65,
        "",
        "[ Three Types of Hours — What Each One Does ]",
        "  This calculator uses three distinct hour concepts. Mixing them up",
        "  leads to incorrect costs or unrealistic production estimates.",
        "",
        "  1. PRODUCTIVE HOURS/DAY (default: 8)",
        "     The time equipment is actively running and moving soil.",
        "     Drives: excavation volume per day, truck trips per day, CO2.",
        "     Does NOT drive operator cost or equipment cost.",
        "     Example: 8 productive hrs x 105 CY/hr excavator = 840 CY/day (before cap).",
        "     Example: 8 productive hrs / 1.1 hr cycle = ~7 truck trips/truck/day.",
        "",
        "  2. OPERATOR PAID HOURS/DAY (default: 10)",
        "     Yard-to-yard time — operators are on the clock from when they",
        "     leave the yard to when they return. Longer than productive hours",
        "     due to drive time, pre/post-trip inspections, and yard time.",
        "     Drives: operator cost and OT calculations ONLY.",
        "     Does NOT affect volume, truck trips, or equipment cost.",
        "     Example: 10 paid hrs x 5 days = 50 hrs/week → 10 OT hrs/week at 1.5x.",
        "",
        "  3. EQUIPMENT DAILY RATE",
        "     Heavy equipment (excavators, loaders) is billed at a flat rate",
        "     per day regardless of how many hours it runs. Billed for both",
        "     working days and weather days.",
        "     Does NOT depend on productive or paid hours at all.",
        "",
        f"  This project: {productive_hours_per_day} productive hrs | "
        f"{operator_paid_hours_per_day} operator paid hrs | "
        f"daily equipment rate",
        "",
        "[ Operator Overtime — 40-Hr Weekly Threshold ]",
        "  Operators are paid yard-to-yard, not just productive field hours.",
        "  Any hours worked beyond 40 in a calendar week are OT at 1.5x rate.",
        "  Weekly paid hrs = Operator Paid Hrs/Day × Work Days/Week",
        "  Regular hrs/week = MIN(40, weekly paid hrs)",
        "  OT hrs/week      = MAX(0, weekly paid hrs - 40)",
        "  Project is split into complete weeks + a partial-week remainder.",
        "  Partial week: hours accumulate day by day; OT kicks in after hr 40.",
        f"  This project: {operator_paid_hours_per_day} hrs/day × {work_days_per_week} days "
        f"= {weekly_paid_hours} hrs/week",
        f"  → {regular_hours_per_week:.0f} regular + {ot_hours_per_week:.0f} OT hrs/week",
        f"  {results['complete_weeks']} complete weeks + {results['remaining_days']} remaining days",
        f"  Partial week: {results['remaining_days']} days × {operator_paid_hours_per_day} hrs "
        f"= {partial_hours} hrs → {partial_regular_hours:.0f} reg + {partial_ot_hours:.0f} OT",
        f"  Project totals: {results['total_operator_regular_hrs']:.0f} regular hrs + "
        f"{results['total_operator_ot_hrs']:.0f} OT hrs",
        "",
        "[ Weather Days & Calendar Duration ]",
        "  Inclement weather days extend the calendar but do not add working days.",
        "  Equipment daily rate IS billed on weather days (equipment sits on site).",
        "  Operators are NOT paid on weather days.",
        "  Crew truck is NOT billed on weather days.",
        "  Calendar days account for both weekends and weather days:",
        "    Calendar Days = (Complete Weeks × 7) + Remaining Days + Weather Days",
        f"  This project: ({results['complete_weeks']} weeks × 7) + "
        f"{results['remaining_days']} days + {weather_days} weather = "
        f"{results['calendar_days']} calendar days",
        f"  Equipment billed for {total_billed_equipment_days} days total.",
        "",
        "[ Mob / Demob ]",
        "  Each piece of heavy equipment incurs a mobilization charge (delivery)",
        "  and a demobilization charge (pickup). Costs are assessed separately",
        "  for excavators and loaders, and billed at 2× the per-unit rate.",
        f"  Excavators: {num_excavators} units × ${excavator_mob_rate:,} × 2 = "
        f"${results['excavator_mob_cost']:,.2f}",
        f"  Loaders:    {num_loaders} units × ${loader_mob_rate:,} × 2 = "
        f"${results['loader_mob_cost']:,.2f}",
        "",
        "[ Energy Surcharge ]",
        f"  {energy_surcharge_pct}% applied to heavy equipment cost only.",
        f"  ${results['total_equipment_cost']:,.2f} × {energy_surcharge_pct}% = "
        f"${results['energy_surcharge_cost']:,.2f}",
        "",
        "[ Environmental Compliance & Insurance (EC&I) ]",
        f"  {eci_pct}% applied to the sum of heavy equipment + operators + misc equipment.",
        f"  Base: ${results['total_equipment_cost']:,.2f} + ${results['total_operator_cost']:,.2f} "
        f"+ ${results['total_misc_cost']:,.2f} = ${eci_base:,.2f}",
        f"  EC&I: ${eci_base:,.2f} × {eci_pct}% = ${results['eci_cost']:,.2f}",
        "",
        "[ Truck Cycle Time ]",
        "  Cycle Time = Loading + Travel to Landfill + Time at Landfill",
        "               + Return Travel + Loading (backfill trip)",
        f"  This project cycle time: {results['trip_time']:.2f} hrs",
        "",
        "[ Trucking Cost ]",
        "  Trucks operate (make trips) within productive hours only. However,",
        "  like equipment operators, truck drivers are paid for the full",
        "  yard-to-yard day (operator paid hours/day).",
        "  Cost = Num Trucks × Operator Paid Hrs/Day × Truck Hourly Rate × Project Days",
        "  Trip utilization % shows what fraction of productive hours are spent",
        "  actively on trips — trips happen in productive hours, pay covers paid hours.",
        f"  This project: {num_trucks} trucks × {operator_paid_hours_per_day} paid hrs × "
        f"${truck_hourly_rate}/hr × {results['project_days']} days = "
        f"${results['trucking_cost']:,.2f}",
        f"  Active trip hours (productive): {results['active_truck_hours']:.0f} hrs "
        f"({results['truck_utilization_pct']:.0f}% trip utilization of productive hours)",
        "",
        "[ Excavation Capacity & Daily Volume Cap ]",
        "  Net capacity = MIN(excavator CY/hr, loader CY/hr) × machine count",
        "  Theoretical daily = net capacity × productive hours",
        "  Cap = max_volume_per_pair × num_pairs",
        "  Effective = MIN(theoretical, cap)",
        f"  Effective: {results['excavation_volume_per_day']:.0f} CY/day "
        f"({'cap binding' if results['volume_cap_active'] else 'theoretical governs'})",
        "",
        "[ Project Duration ]",
        "  Working Days = CEILING(Total Volume / MIN(excavation, trucking) per day)",
        "  Calendar Days = (Complete Weeks × 7) + Remaining Days + Weather Days",
        "  Working days count only days the crew is on site.",
        "  Calendar days reflect the full elapsed duration including weekends and weather.",
        f"  CEILING({total_volume:,} / {limiting_volume:.0f}) = "
        f"{results['project_days']} working days",
        f"  ({results['complete_weeks']} × 7) + {results['remaining_days']} + "
        f"{weather_days} weather = {results['calendar_days']} calendar days",
        "",
        "[ CO2 Emissions ]",
        "  Uses productive hours only — equipment not burning fuel on weather/idle days.",
        "  Productive Project Hours = Working Days × Productive Hrs/Day",
        "  Total Fuel = (Excavators × gal/hr × productive hrs)",
        "             + (Loaders × gal/hr × productive hrs)",
        "             + (Total Truck Hours × truck gal/hr)",
        "  CO2 = Total Fuel × 22.4 lbs/gallon (EPA); tons = lbs / 2,000",
        f"  {results['total_fuel_gallons']:,.0f} gal × 22.4 = "
        f"{co2_lbs:,.0f} lbs = {results['co2_tons']:.2f} tons CO2",
        "",
        "=" * 65,
        "  NOTE: Equipment daily rates include fuel for the machine.",
        "  Fuel inputs are used for CO2 tracking only.",
        "=" * 65,
    ]

    report_text = "\n".join(report_lines)

    # ── Logo preprocessor (removes black background) ─────────────────────────
    def _prep_logo(src_path):
        try:
            from PIL import Image as PILImage
            import numpy as np, tempfile
            img = PILImage.open(src_path).convert("RGBA")
            data = np.array(img)
            mask = (data[:,:,0] < 40) & (data[:,:,1] < 40) & (data[:,:,2] < 40)
            data[mask] = [0, 0, 0, 0]
            result = PILImage.fromarray(data)
            tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
            result.save(tmp.name)
            return tmp.name
        except Exception:
            return src_path

    # ── PDF Quote builder ─────────────────────────────────────────────────────
    def _pdf_build(inp, res, logo_file=None):
        """Build a 3-page professional customer quote PDF. Returns bytes."""
        _buf = io.BytesIO()
        W, H = letter

        # Resolve and preprocess logo (transparent background)
        _raw_logo = None
        for _cand in [logo_file, "Clean_Futures_Cropped.png", "Clean_Futures_2.png"]:
            if _cand and Path(_cand).exists():
                _raw_logo = _cand
                break
        _logo = _prep_logo(_raw_logo) if _raw_logo else None

        C_NAVY    = colors.HexColor("#0D2137")
        C_GREEN   = colors.HexColor("#1A6B2F")
        C_ORANGE  = colors.HexColor("#D4580A")
        C_LBLUE   = colors.HexColor("#D6E4F0")
        C_LGREEN  = colors.HexColor("#EBF5EE")
        C_LYELLOW = colors.HexColor("#FFF8E1")
        C_LGRAY   = colors.HexColor("#F5F5F5")
        C_MGRAY   = colors.HexColor("#E0E0E0")
        C_DGRAY   = colors.HexColor("#555555")

        def _header_footer(canvas, doc):
            canvas.saveState()
            canvas.setFillColor(C_NAVY)
            canvas.rect(0, H - 0.45*inch, W, 0.45*inch, fill=1, stroke=0)
            canvas.setFillColor(C_ORANGE)
            canvas.rect(0, H - 0.52*inch, W, 0.07*inch, fill=1, stroke=0)
            # Logo in header bar
            if _logo:
                try:
                    logo_w, logo_h = 1.15*inch, 0.33*inch
                    canvas.drawImage(_logo, 0.4*inch,
                                     H - 0.45*inch + (0.45*inch - logo_h)/2,
                                     width=logo_w, height=logo_h,
                                     preserveAspectRatio=True, mask='auto')
                except Exception:
                    canvas.setFillColor(colors.white)
                    canvas.setFont("Helvetica-Bold", 9)
                    canvas.drawString(0.5*inch, H - 0.27*inch, "CLEAN FUTURES")
            else:
                canvas.setFillColor(colors.white)
                canvas.setFont("Helvetica-Bold", 9)
                canvas.drawString(0.5*inch, H - 0.27*inch, "CLEAN FUTURES")
            canvas.setFillColor(colors.white)
            canvas.setFont("Helvetica", 8)
            canvas.drawRightString(W - 0.5*inch, H - 0.27*inch,
                                   "Excavation & Remediation Cost Estimate")
            canvas.setFillColor(C_NAVY)
            canvas.rect(0, 0.4*inch, W, 0.025*inch, fill=1, stroke=0)
            canvas.setFillColor(C_DGRAY)
            canvas.setFont("Helvetica", 7)
            canvas.drawString(0.5*inch, 0.22*inch,
                f"Prepared: {date.today().strftime('%B %d, %Y')}  |  "
                "This estimate is based on provided project parameters and is subject to "
                "field conditions and final contract terms.")
            canvas.drawRightString(W - 0.5*inch, 0.22*inch, f"Page {doc.page}")
            canvas.restoreState()

        doc = SimpleDocTemplate(_buf, pagesize=letter,
            leftMargin=0.55*inch, rightMargin=0.55*inch,
            topMargin=1.05*inch, bottomMargin=0.65*inch)

        SS = getSampleStyleSheet()
        def sty(name, parent="Normal", **kw):
            return ParagraphStyle(name, parent=SS[parent], **kw)

        S_SECHEAD = sty("SH","Normal", fontSize=12, fontName="Helvetica-Bold",
                        textColor=colors.white, spaceAfter=0, spaceBefore=10)
        S_BODY    = sty("BD","Normal", fontSize=9, fontName="Helvetica",
                        textColor=colors.black, spaceAfter=4, leading=13)
        S_BODYSM  = sty("BS","Normal", fontSize=8, fontName="Helvetica",
                        textColor=C_DGRAY, spaceAfter=3, leading=11)
        S_NOTE    = sty("NT","Normal", fontSize=8, fontName="Helvetica-Oblique",
                        textColor=C_DGRAY, spaceAfter=4, leading=11)
        S_BOLD    = sty("BL","Normal", fontSize=9, fontName="Helvetica-Bold",
                        textColor=colors.black)

        full_w = W - 1.1*inch
        half   = (full_w - 0.2*inch) / 2

        def section_header(title):
            t = Table([[Paragraph(f"  {title}", S_SECHEAD)]],
                      colWidths=[full_w], rowHeights=[0.30*inch])
            t.setStyle(TableStyle([
                ("BACKGROUND",    (0,0), (-1,-1), C_NAVY),
                ("TOPPADDING",    (0,0), (-1,-1), 6),
                ("BOTTOMPADDING", (0,0), (-1,-1), 6),
                ("LEFTPADDING",   (0,0), (-1,-1), 6),
                ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
            ]))
            return t

        def sub_header(title, bg=C_LBLUE):
            p = Paragraph(f"  <b>{title}</b>",
                sty(f"sub_{title[:6]}","Normal", fontSize=9,
                    fontName="Helvetica-Bold", textColor=colors.black))
            t = Table([[p]], colWidths=[full_w])
            t.setStyle(TableStyle([
                ("BACKGROUND",    (0,0), (-1,-1), bg),
                ("TOPPADDING",    (0,0), (-1,-1), 3),
                ("BOTTOMPADDING", (0,0), (-1,-1), 3),
                ("LEFTPADDING",   (0,0), (-1,-1), 6),
            ]))
            return t

        def fmt(v):  return f"${v:,.0f}"
        def fmtd(v): return f"${v:,.2f}"

        # KPI boxes use plain Paragraphs (no nested tables) to avoid ReportLab layout collisions
        _kpi_center = sty("KC","Normal", alignment=TA_CENTER, leading=16)

        def kpi_cell(label, value, sub_txt, fg=C_NAVY):
            return Paragraph(
                f"<font size='7' color='#555555'>{label}</font><br/>"
                f"<b><font size='16' color='{fg.hexval()}'>{value}</font></b><br/>"
                f"<font size='7' color='#999999'>{sub_txt}</font>",
                _kpi_center
            )

        def simple_tbl(rows, cw, label_col=0):
            t = Table(rows, colWidths=cw)
            style = [
                ("FONTNAME",      (0,0), (-1,-1), "Helvetica"),
                ("FONTSIZE",      (0,0), (-1,-1), 8.5),
                ("FONTNAME",      (label_col,0), (label_col,-1), "Helvetica-Bold"),
                ("TEXTCOLOR",     (label_col,0), (label_col,-1), C_NAVY),
                ("TOPPADDING",    (0,0), (-1,-1), 3),
                ("BOTTOMPADDING", (0,0), (-1,-1), 3),
                ("LEFTPADDING",   (0,0), (-1,-1), 6),
                ("RIGHTPADDING",  (0,0), (-1,-1), 6),
                ("ROWBACKGROUNDS",(0,0), (-1,-1), [colors.white, C_LGRAY]),
                ("GRID",          (0,0), (-1,-1), 0.3, C_MGRAY),
                ("ALIGN",         (1,0), (-1,-1), "RIGHT"),
            ]
            t.setStyle(TableStyle(style))
            return t

        def cost_section(title, line_items, total_label, total_val, note=None, bg=C_LBLUE):
            rows = []
            for label, val in line_items:
                is_sub = label.startswith("  ")
                fsz = 8 if is_sub else 9
                fn  = "Helvetica" if is_sub else "Helvetica-Bold"
                clr = "#666666" if is_sub else "#000000"
                rows.append([
                    Paragraph(f"<font color='{clr}' size='{fsz}'>{label.strip()}</font>",
                        sty(f"cl{id(label)[:4] if hasattr(id(label),'__getitem__') else '0'}","Normal",
                            fontSize=fsz, fontName=fn, leading=11)),
                    Paragraph(f"<font color='{clr}' size='{fsz}'>{val}</font>",
                        sty(f"cv{id(val)[:4] if hasattr(id(val),'__getitem__') else '0'}","Normal",
                            fontSize=fsz, fontName=fn, alignment=TA_RIGHT)),
                ])
            rows.append([
                Paragraph(f"<b>{total_label}</b>",
                    sty("tl","Normal",fontSize=9,fontName="Helvetica-Bold",textColor=colors.white)),
                Paragraph(f"<b>{total_val}</b>",
                    sty("tv","Normal",fontSize=9,fontName="Helvetica-Bold",
                        textColor=colors.white, alignment=TA_RIGHT)),
            ])
            tbl = Table(rows, colWidths=[full_w - 1.6*inch, 1.6*inch])
            style = [
                ("TOPPADDING",    (0,0), (-1,-1), 3),
                ("BOTTOMPADDING", (0,0), (-1,-1), 3),
                ("LEFTPADDING",   (0,0), (-1,-1), 8),
                ("RIGHTPADDING",  (0,0), (-1,-1), 8),
                ("LINEBELOW",     (0,0), (-1,-2), 0.3, C_MGRAY),
                ("BACKGROUND",    (0,-1), (-1,-1), C_NAVY),
                ("TOPPADDING",    (0,-1), (-1,-1), 5),
                ("BOTTOMPADDING", (0,-1), (-1,-1), 5),
            ]
            for i in range(0, len(rows)-1, 2):
                style.append(("BACKGROUND", (0,i), (-1,i), C_LGRAY))
            tbl.setStyle(TableStyle(style))
            block = [sub_header(title, bg), tbl, Spacer(1, 5)]
            if note:
                block.append(Paragraph(note, S_NOTE))
            return KeepTogether(block)

        story = []

        # ── PAGE 1: Cover + Results Summary + Scope ───────────────────────────
        if _logo:
            logo_cell = Image(_logo, width=1.6*inch, height=0.55*inch)
        else:
            logo_cell = Paragraph("<b>CLEAN FUTURES</b>",
                sty("lc","Normal",fontSize=15,fontName="Helvetica-Bold",textColor=C_NAVY))

        title_cell = Paragraph("Excavation &amp; Haul<br/><b>Cost Estimate</b>",
            sty("tc","Normal",fontSize=18,fontName="Helvetica-Bold",textColor=C_NAVY,leading=22))
        date_cell  = Paragraph(
            f"<font color='#555555' size='9'>Prepared: "
            f"{date.today().strftime('%B %d, %Y')}<br/>"
            f"Total Volume: <b>{inp['total_volume']:,} CY</b></font>",
            sty("dc","Normal",alignment=TA_RIGHT,leading=13))

        hdr_tbl = Table([[logo_cell, title_cell, date_cell]],
                        colWidths=[1.8*inch, 3.5*inch, 2.1*inch])
        hdr_tbl.setStyle(TableStyle([
            ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
            ("ALIGN",         (2,0), (2,0),   "RIGHT"),
            ("LEFTPADDING",   (0,0), (-1,-1), 0),
            ("RIGHTPADDING",  (0,0), (-1,-1), 0),
            ("TOPPADDING",    (0,0), (-1,-1), 0),
            ("BOTTOMPADDING", (0,0), (-1,-1), 0),
        ]))
        story.append(hdr_tbl)
        story.append(HRFlowable(width="100%", thickness=2, color=C_ORANGE, spaceAfter=10))

        # KPI summary — 2 rows of 4 boxes, full page width, flat table (no nested tables)
        story.append(section_header("RESULTS SUMMARY"))
        story.append(Spacer(1, 5))

        _kw = (W - 1.1*inch) / 4   # ~1.85" each — fills full content width

        opt_trucks = max(1, round(res['excavation_volume_per_day'] /
                         max(res['trips_per_truck_per_day'] * inp['truck_capacity'], 1)))

        kpi_tbl = Table([
            [kpi_cell("TOTAL PROJECT COST", fmt(res['total_cost']),   "estimated",     C_GREEN),
             kpi_cell("COST PER CY",        fmtd(res['cost_per_cy']), "$/cubic yard"),
             kpi_cell("WORKING DAYS",       str(res['project_days']), "project days"),
             kpi_cell("CALENDAR DAYS",      str(res['calendar_days']),"incl. weekends")],
            [kpi_cell("TOTAL TRUCK TRIPS",  f"{res['num_trips']:,}",  "round trips"),
             kpi_cell("CO\u2082 EMISSIONS", f"{res['co2_tons']:.1f} tons","estimated"),
             kpi_cell("SYSTEM BOTTLENECK",  res['bottleneck'],         "limiting factor", C_ORANGE),
             kpi_cell("OPTIMAL TRUCKS",     str(opt_trucks),           f"vs {inp['num_trucks']} entered")],
        ], colWidths=[_kw]*4, rowHeights=[0.88*inch, 0.88*inch])
        kpi_tbl.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,-1), C_LBLUE),    # all cells blue by default
            ("BACKGROUND",    (0,0), (0,0),   C_LGREEN),   # top-left: total cost green
            ("BACKGROUND",    (2,1), (2,1),   C_LYELLOW),  # bottleneck yellow
            ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
            ("ALIGN",         (0,0), (-1,-1), "CENTER"),
            ("TOPPADDING",    (0,0), (-1,-1), 10),
            ("BOTTOMPADDING", (0,0), (-1,-1), 10),
            ("LEFTPADDING",   (0,0), (-1,-1), 4),
            ("RIGHTPADDING",  (0,0), (-1,-1), 4),
            ("LINEAFTER",     (0,0), (2,1),   0.5, C_MGRAY),
            ("LINEBELOW",     (0,0), (-1,0),  0.5, C_MGRAY),
            ("BOX",           (0,0), (-1,-1), 0.5, C_MGRAY),
        ]))
        story.append(kpi_tbl)
        story.append(Spacer(1, 12))

        # Project scope table
        story.append(section_header("PROJECT SCOPE"))
        story.append(Spacer(1, 4))
        bv = inp['total_volume'] * inp['backfill_pct'] / 100
        scope_rows = [
            ["Excavated Volume",   f"{inp['total_volume']:,} CY",
             "Backfill Volume",    f"{bv:,.0f} CY"],
            ["Productive Hrs/Day", f"{inp['productive_hours_per_day']} hrs",
             "Paid Hrs/Day",       f"{inp['operator_paid_hours_per_day']} hrs (yard-to-yard)"],
            ["Work Days/Week",     f"{inp['work_days_per_week']} days",
             "Weather Days",       f"{inp['weather_days']} days"],
            ["Excavators",         f"{inp['num_excavators']}  @ {inp['excavator_capacity']} CY/hr",
             "Loaders",            f"{inp['num_loaders']}  @ {inp['loader_capacity']} CY/hr"],
            ["Trucks",             f"{inp['num_trucks']}  x {inp['truck_capacity']} CY",
             "Round-Trip Cycle",   f"{res['trip_time']:.2f} hrs"],
        ]
        scope_tbl = Table(scope_rows,
            colWidths=[1.4*inch, 1.7*inch, 1.5*inch, 2.8*inch])
        scope_tbl.setStyle(TableStyle([
            ("FONTNAME",      (0,0), (-1,-1), "Helvetica"),
            ("FONTSIZE",      (0,0), (-1,-1), 9),
            ("FONTNAME",      (0,0), (0,-1), "Helvetica-Bold"),
            ("FONTNAME",      (2,0), (2,-1), "Helvetica-Bold"),
            ("TEXTCOLOR",     (0,0), (0,-1), C_NAVY),
            ("TEXTCOLOR",     (2,0), (2,-1), C_NAVY),
            ("ROWBACKGROUNDS",(0,0), (-1,-1), [colors.white, C_LGRAY]),
            ("TOPPADDING",    (0,0), (-1,-1), 4),
            ("BOTTOMPADDING", (0,0), (-1,-1), 4),
            ("LEFTPADDING",   (0,0), (-1,-1), 6),
            ("GRID",          (0,0), (-1,-1), 0.3, C_MGRAY),
        ]))
        story.append(scope_tbl)
        story.append(Spacer(1, 4))
        story.append(Paragraph(
            "<i>This estimate is based on the project parameters above. Actual costs may vary "
            "based on site conditions, soil classification, regulatory requirements, and contract terms.</i>",
            S_NOTE))

        # ── PAGE 2: Cost Breakdown ────────────────────────────────────────────
        story.append(PageBreak())
        story.append(section_header("DETAILED COST BREAKDOWN"))
        story.append(Spacer(1, 6))

        bill_days = res['project_days'] + inp['weather_days']
        story.append(cost_section(
            "Heavy Equipment  (daily rate x billing days, including weather days)",
            [("  Excavators", f"{inp['num_excavators']} unit(s) x ${inp['excavator_daily_rate']:,}/day x {bill_days} days = {fmt(res['excavator_equipment_cost'])}"),
             ("  Loaders",    f"{inp['num_loaders']} unit(s) x ${inp['loader_daily_rate']:,}/day x {bill_days} days = {fmt(res['loader_equipment_cost'])}")],
            "Total Heavy Equipment", fmt(res['total_equipment_cost']),
            note="Equipment daily rates are billed for all project days plus inclement weather days, "
                 "as equipment remains mobilized on site regardless of weather.",
        ))

        story.append(cost_section(
            "Mobilization / Demobilization  (per unit x 2)",
            [("  Excavators", f"{inp['num_excavators']} unit(s) x ${inp['excavator_mob_rate']:,} x 2 = {fmt(res['excavator_mob_cost'])}"),
             ("  Loaders",    f"{inp['num_loaders']} unit(s) x ${inp['loader_mob_rate']:,} x 2 = {fmt(res['loader_mob_cost'])}")],
            "Total Mob/Demob", fmt(res['total_mob_cost']),
            note="Each unit is charged once for mobilization and once for demobilization.",
        ))

        ot_hrs = res['total_operator_ot_hrs']
        if ot_hrs > 0:
            ot_note = (f"Overtime applies: {inp['operator_paid_hours_per_day']} paid hrs/day x "
                       f"{inp['work_days_per_week']} days = {res['weekly_paid_hours']:.0f} hrs/week, "
                       f"exceeding the 40-hr OT threshold by {res['ot_hours_per_week']:.0f} hrs/week. "
                       f"Project total: {res['total_operator_regular_hrs']:.0f} regular hrs + "
                       f"{ot_hrs:.0f} OT hrs at 1.5x.")
        else:
            ot_note = (f"Operator hours ({inp['operator_paid_hours_per_day']} paid hrs/day x "
                       f"{inp['work_days_per_week']} days = {res['weekly_paid_hours']:.0f} hrs/week) "
                       "are within the 40-hr threshold; no overtime premium applies.")
        story.append(cost_section(
            "Operators  (hourly pay + 1.5x overtime above 40 hrs/week)",
            [("  Excavator Operators", f"{inp['num_excavators']} operator(s) @ ${inp['excavator_operator_rate']}/hr = {fmt(res['excavator_operator_cost'])}"),
             ("  Loader Operators",    f"{inp['num_loaders']} operator(s) @ ${inp['loader_operator_rate']}/hr = {fmt(res['loader_operator_cost'])}")],
            "Total Operators", fmt(res['total_operator_cost']),
            note=ot_note,
        ))

        misc_lines = [
            ("  Crew Truck(s)",  f"{inp['num_crew_trucks']} x ${inp['crew_truck_daily_rate']:,}/day x {res['project_days']} days = {fmt(res['crew_truck_cost'])}"),
            ("  Porta Potty",    f"${inp['porta_potty_daily_rate']}/day x {res['project_days']} days = {fmt(res['porta_potty_cost'])}"),
            ("  Safety Trailer", f"${inp['safety_trailer_daily_rate']}/day x {res['project_days']} days = {fmt(res['safety_trailer_cost'])}"),
            ("  Dump Trailer",   f"${inp['dump_trailer_daily_rate']}/day x {res['project_days']} days = {fmt(res['dump_trailer_cost'])}"),
            ("  Spotters",       f"{inp['num_spotters']} x ${inp['spotter_daily_rate']}/day x {res['project_days']} days = {fmt(res['spotter_cost'])}"),
            ("  Supervisors",    f"{inp['num_supervisors']} x ${inp['supervisor_daily_rate']}/day x {res['project_days']} days = {fmt(res['supervisor_cost'])}"),
            ("  Per Diems",      f"${inp['per_diem_daily_rate']}/day x {res['project_days']} days = {fmt(res['per_diem_cost'])}"),
        ]
        story.append(cost_section(
            "Miscellaneous Equipment & Personnel  (working days only, not charged on weather days)",
            misc_lines, "Total Miscellaneous", fmt(res['total_misc_cost']),
        ))

        fees_lines = [
            ("  Energy Surcharge", f"{inp['energy_surcharge_pct']:.1f}% x {fmt(res['total_equipment_cost'])} equipment = {fmt(res['energy_surcharge_cost'])}"),
            ("  EC&I Fee",         f"{inp['eci_pct']:.1f}% x {fmt(res['total_equipment_cost']+res['total_operator_cost']+res['total_misc_cost'])} base = {fmt(res['eci_cost'])}"),
        ]
        if res['fuel_surcharge_cost'] > 0:
            fees_lines.append(("  Fuel Surcharge", fmt(res['fuel_surcharge_cost'])))
        fees_total = res['energy_surcharge_cost'] + res['eci_cost'] + res['fuel_surcharge_cost']
        story.append(cost_section(
            "Fees & Surcharges",
            fees_lines, "Total Fees", fmt(fees_total),
            note="EC&I (Environmental Compliance & Insurance) is applied to the combined total of "
                 "heavy equipment, operators, and miscellaneous costs. "
                 "Energy surcharge is applied to heavy equipment cost only.",
        ))

        story.append(cost_section(
            "Trucking  (paid hours/day x hourly rate x working days)",
            [("  Truck Cost",       f"{inp['num_trucks']} trucks x {inp['operator_paid_hours_per_day']} paid hrs/day x ${inp['truck_hourly_rate']}/hr x {res['project_days']} days = {fmt(res['trucking_cost'])}"),
             ("  Trip Utilization", f"{res['truck_utilization_pct']:.1f}% (active trip hours / total contracted hours)")],
            "Total Trucking", fmt(res['trucking_cost']),
            note="Trucks are billed for the full paid hours per day regardless of excavation pace. "
                 "Trip utilization shows the percentage of contracted hours spent actively making trips. "
                 "Consider adjusting truck count if utilization is significantly below 100%.",
        ))

        other_total = (res['total_disposal_cost'] + res['total_backfill_cost'] +
                       res['env_consulting_cost'] + res['site_access_contingency'])
        story.append(cost_section(
            "Disposal, Backfill & Other",
            [("  Disposal",                 f"{inp['total_volume']:,} CY x ${inp['disposal_cost']}/CY = {fmt(res['total_disposal_cost'])}"),
             ("  Backfill Material",        f"{bv:,.0f} CY x ${inp['backfill_cost']}/CY = {fmt(res['total_backfill_cost'])}"),
             ("  Environmental Consulting", f"${inp['env_consulting_rate']:,}/day x {inp['env_consulting_days']} days = {fmt(res['env_consulting_cost'])}"),
             ("  Site Access Contingency",  fmt(res['site_access_contingency']))],
            "Total Disposal, Backfill & Other", fmt(other_total),
        ))

        # Grand total
        story.append(Spacer(1, 4))
        grand_rows = [
            [Paragraph("<b>TOTAL PROJECT COST</b>",
                sty("gt","Normal",fontSize=13,fontName="Helvetica-Bold",textColor=colors.white)),
             Paragraph(f"<b>{fmt(res['total_cost'])}</b>",
                sty("gv","Normal",fontSize=13,fontName="Helvetica-Bold",
                    textColor=colors.white, alignment=TA_RIGHT))],
            [Paragraph("Cost per Cubic Yard",
                sty("cy","Normal",fontSize=9,fontName="Helvetica",textColor=colors.white)),
             Paragraph(fmtd(res['cost_per_cy']),
                sty("cyv","Normal",fontSize=9,textColor=colors.white,alignment=TA_RIGHT))],
        ]
        gtbl = Table(grand_rows, colWidths=[full_w - 1.6*inch, 1.6*inch])
        gtbl.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,-1), C_GREEN),
            ("TOPPADDING",    (0,0), (-1,-1), 8),
            ("BOTTOMPADDING", (0,0), (-1,-1), 8),
            ("LEFTPADDING",   (0,0), (-1,-1), 12),
            ("RIGHTPADDING",  (0,0), (-1,-1), 12),
            ("LINEABOVE",     (0,0), (-1,0), 2, C_ORANGE),
            ("LINEBELOW",     (0,-1),(-1,-1), 1, C_ORANGE),
        ]))
        story.append(gtbl)

        # ── PAGE 3: Capacity Analysis + Environmental Impact ──────────────────
        story.append(PageBreak())
        story.append(section_header("CAPACITY ANALYSIS"))
        story.append(Spacer(1, 6))

        lw = half * 0.58; rw = half - lw
        sched_rows = [
            ["Working Days",         str(res['project_days'])],
            ["Calendar Days",        str(res['calendar_days'])],
            ["Complete Weeks",        str(res['complete_weeks'])],
            ["Remaining Days",        str(res['remaining_days'])],
            ["Weather Days",          str(res['weather_days'])],
            ["Total Truck Trips",     f"{res['num_trips']:,}"],
            ["Trips/Truck/Day",       f"{res['trips_per_truck_per_day']:.2f}"],
            ["Cycle Time/Trip",       f"{res['trip_time']:.2f} hrs"],
            ["Backfill Volume",       f"{bv:,.0f} CY"],
        ]
        cap_rows = [
            ["Excavator Capacity",   f"{res['excavator_capacity']:,.0f} CY/hr"],
            ["Loader Capacity",      f"{res['loader_capacity']:,.0f} CY/hr"],
            ["Net Capacity",         f"{res['excavation_capacity']:,.0f} CY/hr"],
            ["Theoretical Vol/Day",  f"{res['excavation_volume_per_day_uncapped']:,.0f} CY"],
            ["Volume Cap/Day",       f"{res['daily_volume_cap']:,.0f} CY"],
            ["Effective Excav/Day",  f"{res['excavation_volume_per_day']:,.0f} CY"],
            ["Effective Truck/Day",  f"{res['truck_volume_per_day']:,.0f} CY"],
            ["System Bottleneck",    res['bottleneck']],
            ["Truck Utilization",    f"{res['truck_utilization_pct']:.1f}%"],
        ]
        st_tbl = simple_tbl(sched_rows, [lw, rw])
        cp_tbl = simple_tbl(cap_rows,   [lw, rw])

        two_col = Table(
            [[sub_header("Schedule", C_LGREEN), Spacer(0.2*inch,1), sub_header("Equipment Capacity", C_LBLUE)],
             [st_tbl, Spacer(0.2*inch,1), cp_tbl]],
            colWidths=[half, 0.2*inch, half])
        two_col.setStyle(TableStyle([
            ("VALIGN",        (0,0), (-1,-1), "TOP"),
            ("TOPPADDING",    (0,0), (-1,-1), 0),
            ("BOTTOMPADDING", (0,0), (-1,-1), 0),
            ("LEFTPADDING",   (0,0), (-1,-1), 0),
            ("RIGHTPADDING",  (0,0), (-1,-1), 0),
        ]))
        story.append(two_col)
        story.append(Spacer(1, 8))

        story.append(sub_header("How Capacity Is Calculated", C_MGRAY))
        story.append(Spacer(1, 4))
        story.append(Paragraph(
            f"<b>Excavation Net Capacity:</b> The limiting output is the lesser of excavator capacity "
            f"({inp['num_excavators']} x {inp['excavator_capacity']} CY/hr = {res['excavator_capacity']:,.0f} CY/hr) "
            f"and loader capacity ({inp['num_loaders']} x {inp['loader_capacity']} CY/hr = "
            f"{res['loader_capacity']:,.0f} CY/hr). The system cannot move soil faster than the slower machine.",
            S_BODY))
        story.append(Paragraph(
            f"<b>Daily Volume Cap:</b> A maximum of {inp['max_volume_per_pair']:,} CY per equipment pair "
            f"limits theoretical output for very long shifts. With {res['num_pairs']} pair(s), "
            f"the cap is {res['daily_volume_cap']:,.0f} CY/day. "
            f"{'The cap was reached on this project.' if res['volume_cap_active'] else 'The cap was not reached.'}",
            S_BODY))
        story.append(Paragraph(
            f"<b>Trucking Bottleneck Check:</b> {inp['num_trucks']} trucks x "
            f"{res['trips_per_truck_per_day']:.2f} trips/day x {inp['truck_capacity']} CY = "
            f"{res['truck_volume_per_day_theoretical']:,.0f} CY/day theoretical truck capacity. "
            f"System bottleneck: <b>{res['bottleneck']}</b>. "
            f"Truck utilization: {res['truck_utilization_pct']:.1f}% of paid hours on active trips.",
            S_BODY))
        story.append(Paragraph(
            f"<b>Operator Pay &amp; Overtime:</b> Operators are paid for "
            f"{inp['operator_paid_hours_per_day']} hrs/day (yard-to-yard), compared to "
            f"{inp['productive_hours_per_day']} productive hrs. At {inp['work_days_per_week']} days/week = "
            f"{res['weekly_paid_hours']:.0f} paid hrs/week. "
            + (f"<b>Overtime applies</b> — {res['ot_hours_per_week']:.0f} hrs/week above the "
               f"40-hr threshold are paid at 1.5x. Project OT total: {res['total_operator_ot_hrs']:.0f} hrs."
               if res['ot_hours_per_week'] > 0 else
               "Within the 40-hr weekly threshold — no overtime premium."),
            S_BODY))

        story.append(Spacer(1, 10))
        story.append(section_header("ENVIRONMENTAL IMPACT"))
        story.append(Spacer(1, 6))

        prod_hrs = res['project_days'] * inp['productive_hours_per_day']
        equip_fuel = (inp['num_excavators']*inp['excavator_fuel'] +
                      inp['num_loaders']*inp['loader_fuel']) * prod_hrs
        truck_fuel = res['total_truck_hours'] * inp['truck_fuel_rate']

        el_lw = half * 0.60; el_rw = half - el_lw
        env_l = [
            ["Productive Project Hours",  f"{prod_hrs:,.0f} hrs"],
            ["Equipment Fuel",            f"{equip_fuel:,.0f} gallons"],
            ["Truck Fuel",                f"{truck_fuel:,.0f} gallons"],
            ["Total Fuel Consumed",       f"{res['total_fuel_gallons']:,.0f} gallons"],
            ["CO2 Emissions",             f"{res['co2_tons']:.2f} metric tons"],
        ]
        env_r = [
            ["Trees Needed to Offset",    f"{int(res['co2_tons']*16.5):,}  (1-year, EPA)"],
            ["Equivalent Car Miles",       f"{int(res['co2_tons']*2500):,}  miles"],
            ["CO2 per CY Excavated",       f"{res['co2_tons']/max(inp['total_volume'],1)*2000:.2f}  lbs/CY"],
            ["Fuel per Working Day",       f"{res['total_fuel_gallons']/max(res['project_days'],1):,.0f}  gal/day"],
            ["CO2 Conversion Factor",      "22.4 lbs CO2 per gallon diesel (EPA)"],
        ]
        env_l_tbl = simple_tbl(env_l, [el_lw, el_rw])
        env_r_tbl = simple_tbl(env_r, [el_lw, el_rw])
        env_two = Table(
            [[sub_header("Fuel & Emissions", C_LGREEN), Spacer(0.2*inch,1), sub_header("Equivalencies", C_LGREEN)],
             [env_l_tbl, Spacer(0.2*inch,1), env_r_tbl]],
            colWidths=[half, 0.2*inch, half])
        env_two.setStyle(TableStyle([
            ("VALIGN",        (0,0),(-1,-1), "TOP"),
            ("TOPPADDING",    (0,0),(-1,-1), 0),
            ("BOTTOMPADDING", (0,0),(-1,-1), 0),
            ("LEFTPADDING",   (0,0),(-1,-1), 0),
            ("RIGHTPADDING",  (0,0),(-1,-1), 0),
        ]))
        story.append(env_two)
        story.append(Spacer(1, 6))
        story.append(Paragraph(
            f"<b>Methodology:</b> Fuel consumption estimated using entered rates "
            f"({inp['excavator_fuel']} gal/hr excavators, {inp['loader_fuel']} gal/hr loaders, "
            f"{inp['truck_fuel_rate']} gal/hr trucks) applied to operating hours. "
            "CO2 calculated using the EPA factor of 22.4 lbs CO2 per gallon of diesel. "
            "Tree equivalency: ~48 lbs CO2 absorbed per tree per year (EPA). "
            "Car mile equivalency: ~0.89 lbs CO2/mile (EPA average passenger vehicle).",
            S_NOTE))

        # Assumptions footnote
        story.append(Spacer(1, 10))
        story.append(HRFlowable(width="100%", thickness=0.5, color=C_MGRAY))
        story.append(Spacer(1, 4))
        story.append(Paragraph("<b>KEY ASSUMPTIONS</b>", S_BOLD))
        assumptions = [
            f"Equipment daily rates billed for {bill_days} days ({res['project_days']} working + {inp['weather_days']} weather). Operators billed for working days only.",
            f"Trucking: {inp['num_trucks']} trucks x {inp['operator_paid_hours_per_day']} paid hrs/day x ${inp['truck_hourly_rate']}/hr. Billed for full paid day, not just productive hours.",
            f"OT threshold: 40 hrs/week. Hours above 40/week paid at 1.5x base rate.",
            f"Round-trip cycle: {res['trip_time']:.2f} hrs "
            + (f"(2x{inp['loading_time']}h load + 2x{inp['travel_time']}h travel + {inp['landfill_time']}h at landfill + {inp['travel_to_backfill']}h backfill travel + {inp['backfill_loading_time']}h backfill load)"
               if not inp['backfill_at_landfill'] else
               f"(2x{inp['loading_time']}h load + 2x{inp['travel_time']}h travel + {inp['landfill_time']}h at landfill, backfill available at landfill)"),
            f"EC&I ({inp['eci_pct']:.1f}%) on equipment + operators + misc. Energy surcharge ({inp['energy_surcharge_pct']:.1f}%) on heavy equipment only.",
            "Estimate excludes permitting fees, soil characterization, and project management unless noted.",
        ]
        for a in assumptions:
            story.append(Paragraph(f"• {a}", S_BODYSM))

        doc.build(story, onFirstPage=_header_footer, onLaterPages=_header_footer)
        _buf.seek(0)
        return _buf.getvalue()

    # ── Build Excel model with live formulas ──────────────────────────────────
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    def _xl_build(inputs):
        """Build the Excel financial model, populating inputs from current session.
        Row layout (left/input column B):
          5-10: Project params | 13-17: Excavation | 20-25: Loader | 28-31: Trucking
          34-45: Misc equip (incl. spotters B39-40, supervisors B41-42, per diems B43,
                             mob/demob B44-45)
          48-52: Fees (EC&I B48, energy B49, env consult rate B50, days onsite B51,
                       site access B52)
          55-60: Trip times | 63-64: Backfill/disposal | 67-69: Fuel surcharge
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Model"
        FONT_NAME = "Arial"
        BLUE = "0000FF"; BLACK = "000000"
        HEADER_BG = "1F4E79"; HEADER_FG = "FFFFFF"
        SUBHDR_BG = "D6E4F0"; RESULT_BG = "F0F7EE"; TOTAL_BG = "FFF2CC"

        def _fill(c): return PatternFill("solid", fgColor=c)
        def _font(bold=False, color=BLACK, size=10, italic=False):
            return Font(name=FONT_NAME, bold=bold, color=color, size=size, italic=italic)
        def _align(h="left", v="center"):
            return Alignment(horizontal=h, vertical=v)

        def _hdr(row, text, c1=1, c2=3, bg=HEADER_BG, fg=HEADER_FG, sz=11):
            ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
            c = ws.cell(row=row, column=c1, value=text)
            c.font = Font(name=FONT_NAME, bold=True, color=fg, size=sz)
            c.fill = _fill(bg); c.alignment = _align()

        def _sub(row, text, c1=1, c2=3, bg=SUBHDR_BG):
            ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
            c = ws.cell(row=row, column=c1, value=text)
            c.font = _font(bold=True); c.fill = _fill(bg); c.alignment = _align()

        def _inp(row, label, value, unit=""):
            a = ws.cell(row=row, column=1, value=label)
            a.font = _font(); a.alignment = _align()
            b = ws.cell(row=row, column=2, value=value)
            b.font = _font(color=BLUE); b.alignment = _align("right")
            c = ws.cell(row=row, column=3, value=unit)
            c.font = _font(color="666666", size=9, italic=True); c.alignment = _align()
            return b

        def _calc(row, label, formula, unit="", bold=False, bg=None):
            e = ws.cell(row=row, column=5, value=label)
            e.font = _font(bold=bold); e.alignment = _align()
            f = ws.cell(row=row, column=6, value=formula)
            f.font = _font(bold=bold, color=BLACK); f.alignment = _align("right")
            g = ws.cell(row=row, column=7, value=unit)
            g.font = _font(color="666666", size=9, italic=True); g.alignment = _align()
            if bg:
                for col in [5, 6, 7]:
                    ws.cell(row=row, column=col).fill = _fill(bg)
            return f

        # Column widths
        for col, w in [("A",36),("B",16),("C",22),("D",2),("E",36),("F",16),("G",26)]:
            ws.column_dimensions[col].width = w

        # Title
        ws.row_dimensions[1].height = 30
        ws.merge_cells("A1:G1")
        t = ws["A1"]
        t.value = "Clean Futures  |  Dig & Haul Cost Model"
        t.font = Font(name=FONT_NAME, bold=True, size=16, color=HEADER_FG)
        t.fill = _fill("0D2137"); t.alignment = _align("center", "center")
        ws.row_dimensions[2].height = 14
        ws.merge_cells("A2:G2")
        sub = ws["A2"]
        sub.value = f"Generated: {date.today().strftime('%B %d, %Y')}    |    Blue = inputs (editable)    |    Black = formulas (do not edit)"
        sub.font = Font(name=FONT_NAME, size=9, italic=True, color="FFFFFF")
        sub.fill = _fill("1F4E79"); sub.alignment = _align("center", "center")
        ws.row_dimensions[3].height = 6

        # ── INPUTS ──────────────────────────────────────────────────────────
        _hdr(4, "  PROJECT PARAMETERS")
        _inp(5,  "Total Volume to Excavate",         inputs["total_volume"],              "CY")
        _inp(6,  "Backfill Volume (% of excavated)", inputs["backfill_pct"],              "%  (0-100)")
        _inp(7,  "Productive Hours per Day",         inputs["productive_hours_per_day"],  "hrs  (volume & trips)")
        _inp(8,  "Operator Paid Hours per Day",      inputs["operator_paid_hours_per_day"],"hrs  (yard-to-yard)")
        _inp(9,  "Work Days per Week",               inputs["work_days_per_week"],        "days  (5, 6, or 7)")
        _inp(10, "Inclement Weather Days",           inputs["weather_days"],              "days  (equip billed, ops not)")
        ws.row_dimensions[11].height = 6

        _hdr(12, "  EXCAVATION EQUIPMENT")
        _inp(13, "Number of Excavators",             inputs["num_excavators"],            "")
        _inp(14, "Excavator Daily Rate",             inputs["excavator_daily_rate"],      "$/day  (incl. fuel)")
        _inp(15, "Excavator Operator Rate",          inputs["excavator_operator_rate"],   "$/hr")
        _inp(16, "Excavator Production",             inputs["excavator_capacity"],        "CY/hr")
        _inp(17, "Excavator Fuel",                   inputs["excavator_fuel"],            "gal/hr  (CO2)")
        ws.row_dimensions[18].height = 6

        _hdr(19, "  LOADER EQUIPMENT")
        _inp(20, "Number of Loaders",                inputs["num_loaders"],               "")
        _inp(21, "Loader Daily Rate",                inputs["loader_daily_rate"],         "$/day  (incl. fuel)")
        _inp(22, "Loader Operator Rate",             inputs["loader_operator_rate"],      "$/hr")
        _inp(23, "Loader Production",                inputs["loader_capacity"],           "CY/hr")
        _inp(24, "Loader Fuel",                      inputs["loader_fuel"],               "gal/hr  (CO2)")
        _inp(25, "Max Daily Volume per Equip Pair",  inputs["max_volume_per_pair"],       "CY")
        ws.row_dimensions[26].height = 6

        _hdr(27, "  TRUCKING")
        _inp(28, "Number of Trucks",                 inputs["num_trucks"],                "")
        _inp(29, "Truck Capacity",                   inputs["truck_capacity"],            "CY")
        _inp(30, "Truck Hourly Rate",                inputs["truck_hourly_rate"],         "$/hr  (incl. driver & fuel)")
        _inp(31, "Truck Fuel",                       inputs["truck_fuel_rate"],           "gal/hr  (CO2)")
        ws.row_dimensions[32].height = 6

        _hdr(33, "  MISCELLANEOUS EQUIPMENT  (working days only)")
        _inp(34, "Number of Crew Trucks",            inputs["num_crew_trucks"],           "")
        _inp(35, "Crew Truck Daily Rate",            inputs["crew_truck_daily_rate"],     "$/day")
        _inp(36, "Porta Potty",                      inputs["porta_potty_daily_rate"],    "$/day")
        _inp(37, "Safety Trailer",                   inputs["safety_trailer_daily_rate"], "$/day")
        _inp(38, "Dump Trailer",                     inputs["dump_trailer_daily_rate"],   "$/day")
        _inp(39, "Number of Spotters",               inputs["num_spotters"],              "")
        _inp(40, "Spotter Daily Rate",               inputs["spotter_daily_rate"],        "$/day")
        _inp(41, "Number of Supervisors",            inputs["num_supervisors"],           "")
        _inp(42, "Supervisor Daily Rate",            inputs["supervisor_daily_rate"],     "$/day")
        _inp(43, "Per Diems — Flat Daily Rate",      inputs["per_diem_daily_rate"],       "$/day  (all-in crew total)")
        _inp(44, "Excavator Mob/Demob",              inputs["excavator_mob_rate"],        "$/unit  (x2: mob+demob)")
        _inp(45, "Loader Mob/Demob",                 inputs["loader_mob_rate"],           "$/unit  (x2: mob+demob)")
        ws.row_dimensions[46].height = 6

        _hdr(47, "  FEES & CONTINGENCIES")
        b48 = _inp(48, "Environmental Compliance & Insurance", inputs["eci_pct"]/100,    "% of equip+ops+misc")
        b48.number_format = "0.0%"
        b49 = _inp(49, "Energy Surcharge",           inputs["energy_surcharge_pct"]/100, "% of heavy equipment")
        b49.number_format = "0.0%"
        _inp(50, "Environmental Consulting",         inputs["env_consulting_rate"],       "$/day")
        _inp(51, "Env. Consulting — Days Onsite",    inputs["env_consulting_days"],       "days")
        _inp(52, "Site Access Contingency",          inputs["site_access_contingency"],   "$  (flat amount)")
        ws.row_dimensions[53].height = 6

        _hdr(54, "  TRIP TIMES")
        _inp(55, "Loading Time",                     inputs["loading_time"],              "hrs")
        _inp(56, "Travel to Landfill (one-way)",     inputs["travel_time"],               "hrs")
        _inp(57, "Time at Landfill (wait + dump)",   inputs["landfill_time"],             "hrs")
        _inp(58, "Backfill Available at Landfill",   1 if inputs["backfill_at_landfill"] else 0, "1=Yes  0=No")
        _inp(59, "Additional Travel Time for Backfill", inputs["travel_to_backfill"],    "hrs  (added to round-trip)")
        _inp(60, "Backfill Loading Time",            inputs["backfill_loading_time"],     "hrs")
        ws.row_dimensions[61].height = 6

        _hdr(62, "  BACKFILL & DISPOSAL")
        _inp(63, "Backfill Cost",                    inputs["backfill_cost"],             "$/CY")
        _inp(64, "Disposal Cost",                    inputs["disposal_cost"],             "$/CY")
        ws.row_dimensions[65].height = 6

        _hdr(66, "  FUEL SURCHARGE  (optional)")
        _inp(67, "Enable Fuel Surcharge",            1 if inputs["fuel_surcharge_enabled"] else 0, "1=Yes  0=No")
        _inp(68, "Surcharge Amount",                 inputs["fuel_surcharge_amount"],     "$")
        _inp(69, "Surcharge Interval",               inputs["fuel_surcharge_interval"],   "daily / weekly / per-trip")

        # ── CALCULATIONS (right side) ────────────────────────────────────────
        _hdr(4, "  SCHEDULE & CAPACITY", 5, 7)
        # B55=loading, B56=travel landfill, B57=landfill time, B58=backfill toggle,
        # B59=extra backfill travel, B60=backfill loading
        _calc(5,  "Round-Trip Cycle Time",           "=2*B55+2*B56+B57+IF(B58=0,B59+B60,0)", "hrs")
        _calc(6,  "Excavator Total Capacity",        "=B13*B16",  "CY/hr")
        _calc(7,  "Loader Total Capacity",           "=B20*B23",  "CY/hr")
        _calc(8,  "Excavation Capacity (net)",
            "=IF(AND(B13>0,B20>0),MIN(B13*B16,B20*B23),IF(B13>0,B13*B16,B20*B23))", "CY/hr")
        _calc(9,  "Equipment Pairs",
            "=IF(AND(B13>0,B20>0),MIN(B13,B20),IF(B13>0,B13,B20))", "")
        _calc(10, "Theoretical Excavation Vol/Day",  "=F8*B7",    "CY")
        _calc(11, "Daily Volume Cap (all pairs)",    "=F9*B25",   "CY")
        _calc(12, "Effective Excavation Vol/Day",    "=MIN(F10,F11)", "CY")
        _calc(13, "Trips per Truck per Day",         "=ROUND(B7/F5,0)", "trips")
        _calc(14, "Truck Vol/Day (theoretical)",     "=F13*B28*B29", "CY")
        _calc(15, "Effective Truck Vol/Day",         "=MIN(F14,F12)", "CY")
        _calc(16, "System Bottleneck",               '=IF(F14<=F12,"Trucking","Excavation")', "")
        _calc(17, "Limiting Volume/Day",             "=MIN(F12,F14)", "CY")
        _calc(18, "Optimal Truck Count",
            "=IFERROR(CEILING(F12/(F13*B29),1),B28)", "trucks")
        ws.row_dimensions[19].height = 6
        _sub(20, "  Project Duration", 5, 7)
        _calc(21, "Project Working Days",            "=CEILING(B5/F17,1)", "days", bold=True, bg=RESULT_BG)
        _calc(22, "Complete Weeks",                  "=INT(F21/B9)",        "weeks")
        _calc(23, "Remaining Days (partial week)",   "=MOD(F21,B9)",        "days")
        _calc(24, "Calendar Days (incl. wknds + wx)","=F22*7+F23+B10",      "days", bold=True, bg=RESULT_BG)
        _calc(25, "Equipment Billing Days",          "=F21+B10",            "days  (working + weather)")
        _calc(26, "Total Truck Trips",               "=CEILING(B5/B29,1)",  "trips")
        _calc(27, "Backfill Volume",                 "=B5*(B6/100)",        "CY")
        ws.row_dimensions[28].height = 6
        _sub(29, "  Operator OT  (40-hr weekly threshold)", 5, 7)
        _calc(30, "Weekly Paid Hours",               "=B8*B9",              "hrs/week")
        _calc(31, "Regular Hrs per Week",            "=MIN(40,F30)",        "hrs")
        _calc(32, "OT Hrs per Week",                 "=MAX(0,F30-40)",      "hrs  (at 1.5x)")
        _calc(33, "Partial Week Hours",              "=F23*B8",             "hrs")
        _calc(34, "Partial Week Regular Hrs",        "=MIN(40,F33)",        "hrs")
        _calc(35, "Partial Week OT Hrs",             "=MAX(0,F33-40)",      "hrs  (at 1.5x)")
        _calc(36, "Total Regular Hrs (per operator)","=F22*F31+F34",        "hrs")
        _calc(37, "Total OT Hrs (per operator)",     "=F22*F32+F35",        "hrs")
        ws.row_dimensions[38].height = 6

        _hdr(39, "  COST BREAKDOWN", 5, 7)
        _sub(40, "  Heavy Equipment  (daily rate x billing days)", 5, 7)
        _calc(41, "  Excavator Equipment",           "=B13*B14*F25",        "$")
        _calc(42, "  Loader Equipment",              "=B20*B21*F25",        "$")
        _calc(43, "  Total Heavy Equipment",         "=F41+F42",            "$", bold=True, bg=TOTAL_BG)
        # B44=excavator mob, B45=loader mob
        _sub(44, "  Mob / Demob  (per unit x 2)", 5, 7)
        _calc(45, "  Excavator Mob/Demob",           "=B13*B44*2",          "$")
        _calc(46, "  Loader Mob/Demob",              "=B20*B45*2",          "$")
        _calc(47, "  Total Mob/Demob",               "=F45+F46",            "$", bold=True, bg=TOTAL_BG)
        _sub(48, "  Operators  (hourly + 1.5x OT above 40 hrs/wk)", 5, 7)
        _calc(49, "  Excavator Operators",           "=B13*(F36*B15+F37*B15*1.5)", "$")
        _calc(50, "  Loader Operators",              "=B20*(F36*B22+F37*B22*1.5)", "$")
        _calc(51, "  Total Operators",               "=F49+F50",            "$", bold=True, bg=TOTAL_BG)
        _sub(52, "  Misc Equipment  (working days only)", 5, 7)
        _calc(53, "  Crew Truck(s)",                 "=B34*B35*F21",        "$")
        _calc(54, "  Porta Potty",                   "=B36*F21",            "$")
        _calc(55, "  Safety Trailer",                "=B37*F21",            "$")
        _calc(56, "  Dump Trailer",                  "=B38*F21",            "$")
        _calc(57, "  Spotters",                      "=B39*B40*F21",        "$")
        _calc(58, "  Supervisors",                   "=B41*B42*F21",        "$")
        _calc(59, "  Per Diems",                     "=B43*F21",            "$")
        _calc(60, "  Total Misc Equipment",          "=F53+F54+F55+F56+F57+F58+F59", "$", bold=True, bg=TOTAL_BG)
        ws.row_dimensions[61].height = 6
        # B48=EC&I%, B49=energy surcharge%
        _calc(62, "Energy Surcharge",                "=B49*F43",            "$")
        _calc(63, "EC&I Base  (equip + ops + misc)", "=F43+F51+F60",        "$")
        _calc(64, "EC&I Fee",                        "=B48*F63",            "$")
        ws.row_dimensions[65].height = 6
        _sub(66, "  Trucking  (paid hrs/day x rate x working days)", 5, 7)
        _calc(67, "  Trucking Cost",                 "=B28*B8*B30*F21",     "$", bold=True, bg=TOTAL_BG)
        _calc(68, "  Trip Utilization %",
            "=IFERROR(ROUND(F15/B29,0)*F5*F21/(B28*B8*F21),0)", "%")
        ws.row_dimensions[69].height = 6
        # B67=enable fuel surcharge, B68=amount, B69=interval
        _calc(70, "Fuel Surcharge",
            '=IF(B67=0,0,IF(B69="daily",B68*F21,IF(B69="weekly",B68*CEILING(F21/7,1),IF(B69="per-trip",B68*F26,0))))',
            "$")
        ws.row_dimensions[71].height = 6
        # B64=disposal cost, B63=backfill cost
        _calc(72, "Disposal Cost",                   "=B5*B64",             "$")
        _calc(73, "Backfill Material Cost",          "=F27*B63",            "$")
        # B50=env consulting rate, B51=env consulting days onsite, B52=site access
        _calc(74, "Environmental Consulting",        "=B50*B51",            "$ ($/day x days onsite)")
        _calc(75, "Site Access Contingency",         "=B52",                "$")
        ws.row_dimensions[76].height = 6

        # Grand Total row
        for col in [5, 6, 7]:
            ws.cell(row=77, column=col).fill = _fill(TOTAL_BG)
        ws.cell(row=77, column=5, value="TOTAL PROJECT COST").font = Font(name=FONT_NAME, bold=True, size=12)
        ws.cell(row=77, column=5).alignment = _align()
        tf = ws.cell(row=77, column=6, value="=F43+F47+F51+F60+F62+F64+F67+F70+F72+F73+F74+F75")
        tf.font = Font(name=FONT_NAME, bold=True, size=12, color=BLACK)
        tf.number_format = '$#,##0'; tf.alignment = _align("right")
        _calc(78, "Cost per Cubic Yard",             "=F77/B5",             "$/CY", bold=True, bg=TOTAL_BG)
        ws.row_dimensions[79].height = 6

        _hdr(80, "  ENVIRONMENTAL IMPACT", 5, 7, bg="2E7D32")
        _calc(81, "Productive Project Hours",        "=F21*B7",             "hrs")
        _calc(82, "Total Truck Hours (CO2)",         "=F26*F5",             "hrs")
        _calc(83, "Equipment Fuel",                  "=(B13*B17+B20*B24)*F81", "gallons")
        _calc(84, "Truck Fuel",                      "=F82*B31",            "gallons")
        _calc(85, "Total Fuel Consumed",             "=F83+F84",            "gallons", bold=True)
        _calc(86, "CO2 Emissions (lbs)",             "=F85*22.4",           "lbs  (EPA)")
        _calc(87, "CO2 Emissions (tons)",            "=F86/2000",           "tons", bold=True)
        _calc(88, "Equivalent Trees to Offset",      "=INT(F87*16.5)",      "trees  (1 yr, EPA)")
        _calc(89, "Equivalent Car Miles",            "=INT(F87*2500)",      "miles  (avg car)")

        # Number formats
        for r in [41,42,43,45,46,47,49,50,51,53,54,55,56,57,58,59,60,62,63,64,67,70,72,73,74,75,77,78]:
            ws.cell(row=r, column=6).number_format = '$#,##0;($#,##0);"-"'
        ws.cell(row=68, column=6).number_format = "0.0%"
        ws.cell(row=77, column=6).number_format = '$#,##0'
        ws.cell(row=78, column=6).number_format = '$#,##0.00'
        for r in [5,10,11,12,13,14,15,17]:
            ws.cell(row=r, column=6).number_format = '0.00'
        for r in [21,22,23,24,25,26,27,30,31,32,33,34,35,36,37,81,82,83,84,85,86,88,89]:
            ws.cell(row=r, column=6).number_format = '#,##0'
        ws.cell(row=87, column=6).number_format = '0.00'
        ws.cell(row=18, column=6).number_format = '0'
        ws.cell(row=5,  column=2).number_format = '#,##0'
        for r in [14,21,30,35,40,42,43,50,63,64,68]:
            ws.cell(row=r, column=2).number_format = '$#,##0'

        ws.freeze_panes = "A3"

        # Legend
        ws.row_dimensions[91].height = 6
        _hdr(92, "  COLOR KEY", 1, 7, bg="37474F", sz=10)
        legend_rows = [
            (93, "BLUE TEXT",    "Input cells — change these to model different scenarios", BLUE),
            (94, "BLACK TEXT",   "Formula cells — calculated automatically, do not edit", BLACK),
            (95, "YELLOW ROWS",  "Key totals and summary outputs", BLACK),
            (96, "OT Logic",     "Operators earn 1.5x for any hours over 40/week (yard-to-yard pay)", BLACK),
            (97, "Trucking Pay", "Trucks billed at paid hrs/day; trips happen within productive hrs only", BLACK),
            (98, "Equipment",    "Heavy equipment billed for ALL days incl. weather; operators NOT billed on weather days", BLACK),
            (99, "Backfill",     "When Backfill at Landfill = 0, additional travel + loading time added to cycle", BLACK),
            (100,"Env. Consult", "Env. Consulting = $/day rate x Days Onsite (independent of project working days)", BLACK),
        ]
        for row, key, desc, color in legend_rows:
            k = ws.cell(row=row, column=1, value=key)
            k.font = Font(name=FONT_NAME, bold=True, color=color, size=9)
            d = ws.cell(row=row, column=2, value=desc)
            d.font = Font(name=FONT_NAME, size=9, color="333333")
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
            ws.row_dimensions[row].height = 14

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    xl_inputs = {
        "total_volume": total_volume, "backfill_pct": backfill_pct,
        "productive_hours_per_day": productive_hours_per_day,
        "operator_paid_hours_per_day": operator_paid_hours_per_day,
        "work_days_per_week": work_days_per_week, "weather_days": weather_days,
        "num_excavators": num_excavators, "excavator_daily_rate": excavator_daily_rate,
        "excavator_operator_rate": excavator_operator_rate,
        "excavator_capacity": excavator_capacity, "excavator_fuel": excavator_fuel,
        "num_loaders": num_loaders, "loader_daily_rate": loader_daily_rate,
        "loader_operator_rate": loader_operator_rate,
        "loader_capacity": loader_capacity, "loader_fuel": loader_fuel,
        "max_volume_per_pair": max_volume_per_pair,
        "num_trucks": num_trucks, "truck_capacity": truck_capacity,
        "truck_hourly_rate": truck_hourly_rate, "truck_fuel_rate": truck_fuel_rate,
        "num_crew_trucks": num_crew_trucks,
        "crew_truck_daily_rate": crew_truck_daily_rate,
        "porta_potty_daily_rate": porta_potty_daily_rate,
        "safety_trailer_daily_rate": safety_trailer_daily_rate,
        "dump_trailer_daily_rate": dump_trailer_daily_rate,
        "num_spotters": num_spotters, "spotter_daily_rate": spotter_daily_rate,
        "num_supervisors": num_supervisors, "supervisor_daily_rate": supervisor_daily_rate,
        "per_diem_daily_rate": per_diem_daily_rate,
        "excavator_mob_rate": excavator_mob_rate, "loader_mob_rate": loader_mob_rate,
        "eci_pct": eci_pct, "energy_surcharge_pct": energy_surcharge_pct,
        "env_consulting_rate": env_consulting_rate,
        "env_consulting_days": env_consulting_days,
        "site_access_contingency": site_access_contingency,
        "loading_time": loading_time, "travel_time": travel_time,
        "landfill_time": landfill_time, "backfill_at_landfill": backfill_at_landfill,
        "travel_to_backfill": travel_to_backfill,
        "backfill_loading_time": backfill_loading_time,
        "backfill_cost": backfill_cost, "disposal_cost": disposal_cost,
        "fuel_surcharge_enabled": fuel_surcharge_enabled,
        "fuel_surcharge_amount": fuel_surcharge_amount,
        "fuel_surcharge_interval": fuel_surcharge_interval,
    }
    xl_bytes  = _xl_build(xl_inputs)
    _logo_file = next((p for p in ["Clean_Futures_Cropped.png", "Clean_Futures_2.png"] if Path(p).exists()), None)
    pdf_bytes = _pdf_build(xl_inputs, results, logo_file=_logo_file)

    dl_col1, dl_col2, dl_col3 = st.columns(3)
    with dl_col1:
        st.download_button(
            label="📄 Download Full Report (.txt)",
            data=report_text,
            file_name="dig_and_haul_report.txt",
            mime="text/plain",
            help="All assumptions + results — paste into AI chat for estimating assistance")
        st.caption("Includes full methodology. Ideal for AI chat prompts.")
    with dl_col2:
        st.download_button(
            label="📊 Download Excel Model (.xlsx)",
            data=xl_bytes,
            file_name="dig_and_haul_model.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        st.caption("Live Excel model with formulas — change any blue cell to recalculate.")
    with dl_col3:
        st.download_button(
            label="📋 Download Customer Quote (.pdf)",
            data=pdf_bytes,
            file_name="dig_and_haul_quote.pdf",
            mime="application/pdf")
        st.caption("Professional 3-page quote with cost breakdown, capacity analysis & CO₂ impact.")

# ── Welcome screen ────────────────────────────────────────────────────────────
else:
    st.info("👈 Enter your project parameters in the sidebar and click **Calculate**")
    st.markdown("""
    ### How to Use This Calculator

    1. **Project info** — volume, productive hours, work days/week, weather days
    2. **Equipment** — excavators and loaders with daily rates, operator rates, mob/demob
    3. **Trucking** — number of trucks, capacity, hourly rate, trip times
    4. **Crew & Site** — crew truck count and daily rate
    5. **Fees** — EC&I %, Energy Surcharge %, Environmental Consulting $/CY, contingency
    6. **Backfill & Disposal** — costs per CY
    7. **Click Calculate**

    ### What You'll Get
    - Total project cost and cost per CY
    - Working days vs. calendar days (with weather)
    - Full cost breakdown across all 12 cost categories
    - Bottleneck analysis (excavation vs. trucking)
    - CO2 emissions tracking
    - Downloadable .txt report and .csv summary

    ### Version 3.1 Updates

    ✅ **Equipment charged daily** — separate daily rate for excavators and loaders  
    ✅ **Operators charged hourly** — separate $/hr rate per machine type  
    ✅ **Weekend OT** — 1.5x operator rate on Sat/Sun when working 6 or 7 day weeks  
    ✅ **Mob/Demob** — per-unit charge assessed 2x (mob + demob), separate by type  
    ✅ **Crew truck** — daily charge, not billed on weather days  
    ✅ **Weather days** — extends calendar; equipment still billed, operators/crew not  
    ✅ **Energy Surcharge** — % of heavy equipment cost  
    ✅ **EC&I fee** — % of equipment + operators + crew truck  
    ✅ **Environmental Consulting** — $/CY × total volume  
    ✅ **Site Access Contingency** — flat dollar amount  

    ### Version 3.2 Updates

    ✅ **Operator OT revised** — now based on 40-hr weekly threshold (not weekend detection)  
    ✅ **Operator Paid Hours/Day** — new input (default 10 hrs, yard-to-yard) drives OT math  
    ✅ **Partial week handling** — remaining days at end of project accumulate hours precisely;  
       OT kicks in after the 40th hour within that partial week  
    ✅ **OT always applies** — at 10 hrs/day × 5 days = 50 hrs/week, every week has 10 OT hrs  

    ### Version 3.3 Updates

    ✅ **Hours clarity** — sidebar labels and help text now explicitly explain the  
       three-way split: productive hrs (volume/trips/CO2), operator paid hrs  
       (cost/OT), and daily equipment rate (flat, not hour-based)  
    ✅ **Hours summary caption** — live reminder shown below the hours inputs  
       confirming which value drives which calculation  
    ✅ **Report methodology updated** — Section 3 now has a dedicated "Three Types  
       of Hours" block with worked examples for each  

    ### Version 3.4 Updates

    ✅ **Trucking cost model corrected** — trucks are now billed for all productive  
       hours on site (num trucks × hrs/day × rate × project days). Extra trucks now  
       correctly increase cost even when excavation is the bottleneck  
    ✅ **Truck utilization %** — new metric in Capacity tab and report shows what  
       fraction of contracted truck hours are actively on trips vs. idle  
    ✅ **Bottleneck tip updated** — now recommends optimal truck count when excess  
       trucks are detected  
    ✅ **Default values updated** — total volume, equipment rates, fees, and trip  
       times updated to match current project inputs  

    ### Version 3.5 Updates

    ✅ **Calendar days bug fixed** — was incorrectly showing the same value as  
       working days; now correctly accounts for weekends:  
       Calendar Days = (Complete Weeks × 7) + Remaining Days + Weather Days  
    ✅ **Example:** 162 working days at 5 days/week = 32 full weeks + 2 days  
       = 226 calendar days (not 162)  

    ### Version 3.6 Updates

    ✅ **Trucking cost corrected** — trucks operate within productive hours but are  
       paid for the full yard-to-yard day (operator paid hours). Cost now uses  
       paid hours/day, not productive hours/day  
    ✅ **Truck rate updated** — default hourly rate updated to $105/hr  
    ✅ **Miscellaneous Equipment section** — Crew Truck moved here; added  
       Porta Potty, Safety Trailer, and Dump Trailer as $/day inputs (default $0)  
    ✅ **EC&I base updated** — now includes all misc equipment, not just crew truck  

    ### Version 3.7 Updates

    ✅ **Truck utilization recommendation fixed** — previous formula was circular  
       (used current utilization % × current truck count), causing the suggestion  
       to chase itself lower with each truck removed  
    ✅ **Correct formula:** optimal trucks = CEILING(excavation CY/day ÷ (trips/truck × truck capacity))  
       — independent of how many trucks are currently entered  

    ### Version 3.8 Updates

    ✅ **Mob/Demob moved** — inputs relocated from equipment sections into Miscellaneous  
       Equipment section for cleaner organization  
    ✅ **Mob/Demob defaults** — reduced from $5,000 to $2,500 per unit  
    ✅ **Backfill default** — "Backfill Available at Landfill" now unchecked by default  
    ✅ **Backfill travel renamed** — "Travel to Backfill Site" → "Additional Travel Time  
       for Backfill" to clarify it's extra time added on top of the landfill round-trip  

    ### Version 3.9 Updates

    ✅ **Excel model export** — CSV replaced with a full Excel financial model (.xlsx)  
    ✅ **Live formulas** — every calculated value in the Excel file is a real formula;  
       change any blue input cell and all results recalculate instantly  
    ✅ **Full model structure** — inputs on the left (blue), schedule/capacity/cost/  
       environmental calculations on the right (black), with color-coded sections  
    ✅ **OT, trucking, mob/demob, fuel surcharge** all expressed as Excel formulas  

    ### Version 4.0 Updates

    ✅ **Spotters** — # of Spotters and daily rate added to Miscellaneous Equipment  
    ✅ **Supervisors** — # of Supervisors and daily rate added to Miscellaneous Equipment  
    ✅ **Per Diems** — flat daily rate added to Miscellaneous Equipment (billed × working days)  
    ✅ **Environmental Consulting** — changed from $/CY to $/day with separate Days Onsite  
       input, so the consultant is not assumed to be on site every project day  
    ✅ **Version fix** — title page now uses a single APP_VERSION constant, so it can  
       never fall out of sync with the footer and report header again  

    ### Version 4.4 Updates

    ✅ **Customer Quote PDF** — new 📋 Download button generates a professional 3-page PDF  
    ✅ **Page 1:** Results summary (8 KPI boxes), project scope snapshot  
    ✅ **Page 2:** Full cost breakdown with line-item detail and explanatory notes for each  
       cost category (equipment billing logic, OT rules, trucking pay structure, EC&I base, etc.)  
    ✅ **Page 3:** Capacity analysis (schedule + equipment capacity tables side-by-side),  
       calculation methodology notes, environmental impact tables, key assumptions footnotes  
    ✅ **Brand styling** — Clean Futures navy/green/orange color scheme, logo on every page  
    """)

# ── Footer ────────────────────────────────────────────────────────────────────
st.markdown("---")
footer_col1, footer_col2 = st.columns([3, 1])
with footer_col1:
    st.markdown(f"**Dig and Haul Cost Calculator** v{APP_VERSION} | Built by Clean Futures with Streamlit")
with footer_col2:
    logo_path = Path("Clean_Futures_2.png")
    if logo_path.exists():
        st.image(str(logo_path), width=150)
